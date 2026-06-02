import os
import asyncio
from typing import Optional, List
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import FileResponse, RedirectResponse
from pydantic import BaseModel, Field
from backend.deps import get_current_user, write_access, read_only, get_db, Session
import backend.services.billing_service as bill_svc
import backend.services.property_service as prop_svc
import backend.services.system_service as sys_svc
import backend.services.payment_service as pay_svc
import backend.services.analytics_service as analytics
from backend.generators import soa_gen, computation_gen, notice_gen, report_gen
from utils.logger import mto_logger
from backend.services.storage_service import storage_service

router = APIRouter(tags=["Billing"])


@router.get("/billing/summary")
async def get_billing_summary(current_user: dict = Depends(get_current_user), db_session: Session = Depends(get_db)):
    return sys_svc.get_dashboard_summary(db_session=db_session)

@router.get("/properties/{property_id}/statement")
async def get_property_statement(
    property_id: int, current_user: dict = Depends(get_current_user), db_session: Session = Depends(get_db)
):
    return bill_svc.get_property_statement_data(property_id, db_session=db_session)

@router.get("/billing/assessment-roll")
async def get_assessment_roll(
    limit: int = 100,
    cursor: Optional[int] = None,
    current_user: dict = Depends(get_current_user),
    db_session: Session = Depends(get_db),
):
    return prop_svc.get_assessment_roll(limit=limit, cursor=cursor, db_session=db_session)

@router.get("/billing/report-details")
async def get_report_details(
    month: str = "All",
    year: str = "All",
    limit: int = 200,
    cursor: Optional[int] = None,
    current_user: dict = Depends(get_current_user),
    db_session: Session = Depends(get_db)
):
    return bill_svc.get_report_details(month, year, limit=limit, cursor=cursor, db_session=db_session)

@router.get("/billing/receivables-summary")
async def get_receivables_summary(
    year: str, current_user: dict = Depends(get_current_user), db_session: Session = Depends(get_db)
):
    return bill_svc.get_rpt_receivables_summary(year, db_session=db_session)

@router.get("/billing/delinquents")
async def get_delinquent_list(
    limit: int = 50,
    cursor: Optional[int] = None,
    current_user: dict = Depends(get_current_user),
    db_session: Session = Depends(get_db),
):
    return bill_svc.get_delinquent_accounts(limit=limit, cursor=cursor, db_session=db_session)


@router.get("/billing/collections", dependencies=[Depends(read_only)])
async def get_collections_worklist(
    barangay: Optional[str] = None,
    min_age_days: int = 0,
    limit: int = 50,
    offset: int = 0,
    current_user: dict = Depends(get_current_user),
    db_session: Session = Depends(get_db),
):
    """
    Collections worklist — delinquent accounts prioritised by balance with
    aging buckets (CURRENT / 30 / 60 / 90 / 120+). Drives the staff
    collections dashboard. Returns per-page items plus full-set summary totals.
    """
    return bill_svc.get_collections_worklist(
        barangay=barangay,
        min_age_days=min_age_days,
        limit=limit,
        offset=offset,
        db_session=db_session,
    )


@router.get("/billing/compliant")
async def get_compliant_list(
    barangay: Optional[str] = None,
    limit: int = 50,
    cursor: Optional[int] = None,
    current_user: dict = Depends(get_current_user),
    db_session: Session = Depends(get_db),
):
    """
    Returns properties with zero outstanding balance (fully paid across all years).
    Optionally filtered by barangay. Cursor-paginated.
    """
    return bill_svc.get_compliant_accounts(
        barangay=barangay, limit=limit, cursor=cursor, db_session=db_session
    )


@router.get("/billing/compliant/summary")
async def get_compliant_summary(
    current_user: dict = Depends(get_current_user),
    db_session: Session = Depends(get_db),
):
    """
    Returns per-barangay compliance summary:
    total properties, compliant count, delinquent count, compliance rate %.
    Used for the summary cards at the top of the Compliant Properties dashboard.
    """
    return bill_svc.get_compliant_summary_by_barangay(db_session=db_session)

@router.get("/reports/receivables-by-barangay")
async def get_receivables_by_barangay(
    year: Optional[int] = None,
    data_start_year: int = 2023,
    current_user: dict = Depends(get_current_user),
    db_session: Session = Depends(get_db),
):
    return prop_svc.get_receivables_by_barangay(
        report_year=year,
        data_start_year=data_start_year,
        db_session=db_session,
    )


@router.get("/reports/receivables-by-barangay-pdf", tags=["Reports"])
async def export_receivables_by_barangay_pdf(
    year: Optional[int] = None,
    barangay: Optional[str] = None,
    current_user: dict = Depends(get_current_user),
    db_session: Session = Depends(get_db),
):
    """
    Generates a print-ready PDF of the Receivables by Barangay report
    for the requested year (or all years when year is omitted).
    """
    rows = prop_svc.get_receivables_by_barangay(
        report_year=year, db_session=db_session
    )
    if barangay and barangay != "ALL":
        rows = [r for r in rows if r[0] == barangay]

    year_label = str(year) if year else "All Years"
    if barangay and barangay != "ALL":
        year_label += f" - Barangay: {barangay}"

    try:
        base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        pdf_path = await asyncio.to_thread(
            report_gen.generate_receivables_by_barangay_pdf, rows, year_label, base_dir
        )
        file_name = os.path.basename(pdf_path)

        if storage_service.enabled:
            s3_key = f"reports/{file_name}"
            uploaded_key = await asyncio.to_thread(storage_service.upload_file, pdf_path, s3_key)
            if uploaded_key:
                presigned_url = await asyncio.to_thread(storage_service.generate_presigned_url, s3_key)
                if presigned_url:
                    try:
                        os.remove(pdf_path)
                    except Exception:
                        pass
                    return RedirectResponse(presigned_url, status_code=307)

        return FileResponse(pdf_path, media_type="application/pdf", filename=file_name)
    except Exception as e:
        import traceback
        mto_logger.error(f"Failed to generate Receivables by Barangay PDF: {e}\n{traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"PDF Generation Failed: {str(e)}")


@router.get("/reports/assessment-roll-pdf", tags=["Reports"])
async def export_assessment_roll_pdf(
    barangay: Optional[str] = None,
    year_start: Optional[int] = None,
    year_end: Optional[int] = None,
    current_user: dict = Depends(get_current_user),
    db_session: Session = Depends(get_db),
):
    """
    Generates a print-ready PDF of the full Assessment Roll (landscape A4).
    Optionally filtered by barangay and/or effectivity year range.
    """
    result = prop_svc.search_properties(
        "",
        limit=10000,
        barangay=barangay if barangay and barangay != "ALL" else None,
        year_start=year_start,
        year_end=year_end,
        db_session=db_session,
    )
    items = result.get("items", []) if isinstance(result, dict) else result
    try:
        base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        pdf_path = await asyncio.to_thread(
            report_gen.generate_assessment_roll_pdf, items, base_dir, barangay
        )
        file_name = os.path.basename(pdf_path)

        if storage_service.enabled:
            s3_key = f"reports/{file_name}"
            uploaded_key = await asyncio.to_thread(storage_service.upload_file, pdf_path, s3_key)
            if uploaded_key:
                presigned_url = await asyncio.to_thread(storage_service.generate_presigned_url, s3_key)
                if presigned_url:
                    try:
                        os.remove(pdf_path)
                    except Exception:
                        pass
                    return RedirectResponse(presigned_url, status_code=307)

        return FileResponse(pdf_path, media_type="application/pdf", filename=file_name)
    except Exception as e:
        import traceback
        mto_logger.error(f"Failed to generate Assessment Roll PDF: {e}\n{traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"PDF Generation Failed: {str(e)}")


@router.get("/analytics/trends", tags=["Analytics"], dependencies=[Depends(read_only)])
async def get_analytics_trends(months: int = 12, current_user: dict = Depends(get_current_user), db_session: Session = Depends(get_db)):
    return pay_svc.get_monthly_collection_trend(months, db_session=db_session)

@router.get("/analytics/barangay-breakdown", tags=["Analytics"], dependencies=[Depends(read_only)])
async def get_barangay_breakdown(current_user: dict = Depends(get_current_user), db_session: Session = Depends(get_db)):
    return pay_svc.get_revenue_by_barangay(db_session=db_session)

@router.get("/analytics/kpis", tags=["Analytics"], dependencies=[Depends(read_only)])
async def get_analytics_kpis(current_user: dict = Depends(get_current_user), db_session: Session = Depends(get_db)):
    return pay_svc.get_collection_kpis(db_session=db_session)

@router.get("/api/analytics/dashboard")
async def get_analytics_dashboard(user: str = Depends(get_current_user), db_session: Session = Depends(get_db)):
    """Returns a comprehensive set of treasury analytics data including year-over-year comparison."""
    return {
        "summary":   analytics.get_collection_summary(db_session=db_session),
        "last_year": analytics.get_last_year_summary(db_session=db_session),
        "trend":     analytics.get_monthly_revenue_trend(db_session=db_session),
        "barangays": analytics.get_barangay_distribution(db_session=db_session),
        "years":     analytics.get_tax_year_distribution(db_session=db_session),
    }





@router.get("/properties/{property_id}/computation-pdf", tags=["Financial"])
async def generate_computation_pdf(
    property_id: int, current_user: dict = Depends(get_current_user), db_session: Session = Depends(get_db)
):
    try:
        details = bill_svc.get_property_statement_data(property_id, db_session=db_session)
        if not details:
            raise HTTPException(status_code=404, detail="Property billing data not found")

        base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        pdf_path = await asyncio.to_thread(
            computation_gen.generate_delinquency_computation, details, base_dir
        )
        file_name = os.path.basename(pdf_path)

        if storage_service.enabled:
            s3_key = f"computations/{file_name}"
            uploaded_key = await asyncio.to_thread(storage_service.upload_file, pdf_path, s3_key)
            if uploaded_key:
                presigned_url = await asyncio.to_thread(storage_service.generate_presigned_url, s3_key)
                if presigned_url:
                    try:
                        os.remove(pdf_path)
                    except Exception as cleanup_err:
                        mto_logger.warning(f"Failed to remove local temp PDF '{pdf_path}': {cleanup_err}")
                    return RedirectResponse(presigned_url, status_code=307)

        return FileResponse(pdf_path, media_type="application/pdf", filename=file_name)
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        mto_logger.error(f"Failed to generate computation PDF for property {property_id} | Error: {str(e)}\n{traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"PDF Generation Failed: {str(e)}")

@router.get("/properties/{property_id}/statement-pdf", tags=["Financial"])
async def generate_statement_pdf(
    property_id: int, current_user: dict = Depends(get_current_user), db_session: Session = Depends(get_db)
):
    try:
        details = bill_svc.get_property_statement_data(property_id, db_session=db_session)
        if not details:
            raise HTTPException(status_code=404, detail="Property billing data not found")

        base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        pdf_path = await asyncio.to_thread(
            soa_gen.generate_statement_of_account, details, base_dir
        )
        file_name = os.path.basename(pdf_path)

        if storage_service.enabled:
            s3_key = f"statements/{file_name}"
            uploaded_key = await asyncio.to_thread(storage_service.upload_file, pdf_path, s3_key)
            if uploaded_key:
                presigned_url = await asyncio.to_thread(storage_service.generate_presigned_url, s3_key)
                if presigned_url:
                    try:
                        os.remove(pdf_path)
                    except Exception as cleanup_err:
                        mto_logger.warning(f"Failed to remove local temp PDF '{pdf_path}': {cleanup_err}")
                    return RedirectResponse(presigned_url, status_code=307)

        return FileResponse(pdf_path, media_type="application/pdf", filename=file_name)
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        mto_logger.error(f"Failed to generate SOA PDF for property {property_id} | Error: {str(e)}\n{traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"PDF Generation Failed: {str(e)}")

class BulkSOARequest(BaseModel):
    property_ids: List[int] = Field(..., min_length=1, max_length=200)


@router.post("/billing/bulk-soa-pdf", tags=["Financial"])
async def generate_bulk_soa_pdf(
    data: BulkSOARequest,
    current_user: dict = Depends(write_access),
    db_session: Session = Depends(get_db),
):
    """
    Generates a single merged PDF containing Statement of Account pages
    for each requested property ID. Returns the file directly or a
    pre-signed S3 URL when cloud storage is enabled.
    """
    statements = []
    missing_ids = []

    for pid in data.property_ids:
        details = bill_svc.get_property_statement_data(pid, db_session=db_session)
        if details:
            statements.append(details)
        else:
            missing_ids.append(pid)

    if missing_ids:
        mto_logger.warning(
            f"Bulk SOA: {len(missing_ids)} property ID(s) not found — skipped",
            missing=missing_ids,
            user=current_user.get("username"),
        )

    if not statements:
        raise HTTPException(
            status_code=404,
            detail="No valid property billing data found for the provided IDs.",
        )

    try:
        base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        pdf_path = await asyncio.to_thread(soa_gen.bulk_generate_soa, statements, base_dir)
        file_name = os.path.basename(pdf_path)

        mto_logger.info(
            f"Bulk SOA generated: {file_name} ({len(statements)} properties)",
            user=current_user.get("username"),
        )

        if storage_service.enabled:
            s3_key = f"statements/{file_name}"
            uploaded_key = await asyncio.to_thread(storage_service.upload_file, pdf_path, s3_key)
            if uploaded_key:
                presigned_url = await asyncio.to_thread(storage_service.generate_presigned_url, s3_key)
                if presigned_url:
                    try:
                        os.remove(pdf_path)
                    except Exception as cleanup_err:
                        mto_logger.warning(
                            f"Failed to remove local temp PDF '{pdf_path}': {cleanup_err}"
                        )
                    return RedirectResponse(presigned_url, status_code=307)

        return FileResponse(pdf_path, media_type="application/pdf", filename=file_name)

    except Exception as e:
        import traceback
        mto_logger.error(
            f"Bulk SOA generation failed | Error: {str(e)}\n{traceback.format_exc()}"
        )
        raise HTTPException(
            status_code=500, detail=f"Bulk SOA PDF Generation Failed: {str(e)}"
        )


@router.get("/properties/{property_id}/notice-pdf", tags=["Financial"])
async def generate_notice_pdf(
    property_id: int, current_user: dict = Depends(get_current_user), db_session: Session = Depends(get_db)
):
    try:
        details = bill_svc.get_property_statement_data(property_id, db_session=db_session)
        if not details:
            raise HTTPException(status_code=404, detail="Property billing data not found")

        base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        pdf_path = await asyncio.to_thread(
            notice_gen.generate_delinquency_notice, details, base_dir
        )
        file_name = os.path.basename(pdf_path)

        if storage_service.enabled:
            s3_key = f"notices/{file_name}"
            uploaded_key = await asyncio.to_thread(storage_service.upload_file, pdf_path, s3_key)
            if uploaded_key:
                presigned_url = await asyncio.to_thread(storage_service.generate_presigned_url, s3_key)
                if presigned_url:
                    try:
                        os.remove(pdf_path)
                    except Exception as cleanup_err:
                        mto_logger.warning(f"Failed to remove local temp PDF '{pdf_path}': {cleanup_err}")
                    return RedirectResponse(presigned_url, status_code=307)

        return FileResponse(pdf_path, media_type="application/pdf", filename=file_name)
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        mto_logger.error(f"Failed to generate Notice PDF for property {property_id} | Error: {str(e)}\n{traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"PDF Generation Failed: {str(e)}")


@router.get("/analytics")
async def serve_analytics_dashboard():
    """
    Serves the analytics dashboard HTML from backend/static/analytics.html.
    No HTTP-level auth — the HTML page itself handles authentication by
    reading window.__MTO_TOKEN__ injected by the desktop WebView before load.
    If the token is missing or invalid, the page shows an UNAUTHORIZED message.
    """
    from fastapi.responses import FileResponse
    html_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "static", "analytics.html")
    if not os.path.exists(html_path):
        raise HTTPException(status_code=404, detail="Analytics dashboard file not found.")
    return FileResponse(html_path, media_type="text/html")


# ---------------------------------------------------------------------------
# Excel Export — for COA auditors and management reporting
# ---------------------------------------------------------------------------

class ExportReportRequest(BaseModel):
    month: str = "All"
    year: str = "All"
    report_type: str = "collections"   # collections | delinquents | assessment_roll | receivables
    barangay: Optional[str] = None


@router.post("/billing/export/excel", tags=["Reports"])
async def export_billing_excel(
    data: ExportReportRequest,
    current_user: dict = Depends(read_only),
    db_session: Session = Depends(get_db),
):
    """
    Exports billing/collection data as an Excel (.xlsx) file.

    report_type values:
      collections     — payment report details (filtered by month/year)
      delinquents     — all properties with outstanding balances
      assessment_roll — full assessment roll with billing summary

    Returns the file directly as a download. openpyxl is already in
    requirements.txt so no new dependency is needed.

    COA auditors need Excel, not PDFs — this endpoint closes that gap.
    """
    import io
    import asyncio
    from datetime import datetime, timezone
    from fastapi.responses import StreamingResponse
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    def _build_workbook() -> bytes:
        wb = openpyxl.Workbook()
        ws = wb.active

        # ── Shared styles ────────────────────────────────────────────────────
        header_font  = Font(bold=True, color="FFFFFF", size=11)
        header_fill  = PatternFill("solid", fgColor="1F4E78")
        title_font   = Font(bold=True, size=13)
        center_align = Alignment(horizontal="center", vertical="center")
        right_align  = Alignment(horizontal="right")
        thin_border  = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )
        currency_fmt = '#,##0.00'
        date_str     = datetime.now(timezone.utc).strftime("%B %d, %Y %I:%M %p UTC")

        def style_header_row(ws, row_num: int, col_count: int):
            for col in range(1, col_count + 1):
                cell = ws.cell(row=row_num, column=col)
                cell.font   = header_font
                cell.fill   = header_fill
                cell.alignment = center_align
                cell.border = thin_border

        def auto_width(ws):
            for col in ws.columns:
                max_len = max((len(str(c.value or "")) for c in col), default=10)
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)

        # ── Report: Collections ───────────────────────────────────────────────
        if data.report_type == "collections":
            ws.title = "Collections"
            ws["A1"] = "MUNICIPAL TREASURY OFFICE — COLLECTION REPORT"
            ws["A1"].font = title_font
            ws["A2"] = f"Period: {data.month} / {data.year}    Generated: {date_str}"
            ws["A2"].font = Font(italic=True, size=10)
            ws.merge_cells("A1:I1")
            ws.merge_cells("A2:I2")
            ws["A1"].alignment = center_align
            ws["A2"].alignment = center_align

            headers = ["TD Number", "Owner Name", "Barangay", "Tax Year",
                       "OR Number", "Date Paid", "Basic", "SEF", "Total Paid"]
            for col_idx, h in enumerate(headers, 1):
                ws.cell(row=4, column=col_idx, value=h)
            style_header_row(ws, 4, len(headers))

            rows = bill_svc.get_report_details(
                data.month, data.year, limit=10000, cursor=None, db_session=db_session
            )
            if isinstance(rows, dict) and "items" in rows:
                rows = rows["items"]

            total_basic = total_sef = total_paid = 0.0
            for row_idx, row in enumerate(rows or [], start=5):
                if isinstance(row, dict):
                    vals = [
                        row.get("td_number", ""), row.get("owner_name", ""),
                        row.get("barangay", ""), row.get("tax_year", ""),
                        row.get("or_number", ""), row.get("date_paid", ""),
                        float(row.get("basic_amount", 0) or 0),
                        float(row.get("sef_amount", 0) or 0),
                        float(row.get("amount_paid", 0) or 0),
                    ]
                else:
                    vals = list(row) + [""] * max(0, 9 - len(row))

                for col_idx, val in enumerate(vals, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    cell.border = thin_border
                    if col_idx >= 7:
                        cell.number_format = currency_fmt
                        cell.alignment = right_align

                if len(vals) >= 9:
                    total_basic += float(vals[6] or 0)
                    total_sef   += float(vals[7] or 0)
                    total_paid  += float(vals[8] or 0)

            # Totals row
            total_row = len(rows or []) + 5
            ws.cell(row=total_row, column=6, value="TOTAL").font = Font(bold=True)
            for col_idx, total in [(7, total_basic), (8, total_sef), (9, total_paid)]:
                cell = ws.cell(row=total_row, column=col_idx, value=total)
                cell.font = Font(bold=True)
                cell.number_format = currency_fmt
                cell.alignment = right_align

        # ── Report: Delinquents ───────────────────────────────────────────────
        elif data.report_type == "delinquents":
            ws.title = "Delinquent Accounts"
            ws["A1"] = "MUNICIPAL TREASURY OFFICE — DELINQUENT ACCOUNTS"
            ws["A1"].font = title_font
            ws["A2"] = f"Generated: {date_str}"
            ws["A2"].font = Font(italic=True, size=10)
            ws.merge_cells("A1:G1")
            ws.merge_cells("A2:G2")
            ws["A1"].alignment = center_align

            headers = ["TD Number", "Owner Name", "Location", "Total Due", "Total Paid", "Balance", "Status"]
            for col_idx, h in enumerate(headers, 1):
                ws.cell(row=4, column=col_idx, value=h)
            style_header_row(ws, 4, len(headers))

            result = bill_svc.get_delinquent_accounts(limit=10000, cursor=None, db_session=db_session)
            items = result.get("items", []) if isinstance(result, dict) else result

            total_balance = 0.0
            for row_idx, item in enumerate(items or [], start=5):
                if isinstance(item, dict):
                    balance = float(item.get("balance", 0) or 0)
                    vals = [
                        item.get("td_number", ""), item.get("owner_name", ""),
                        item.get("location", ""),
                        float(item.get("total_due", 0) or 0),
                        float(item.get("total_paid", 0) or 0),
                        balance,
                        "DELINQUENT" if balance > 0 else "SETTLED",
                    ]
                else:
                    vals = list(item) + [""] * max(0, 7 - len(item))

                for col_idx, val in enumerate(vals, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    cell.border = thin_border
                    if col_idx in (4, 5, 6):
                        cell.number_format = currency_fmt
                        cell.alignment = right_align

                total_balance += float(vals[5] or 0)

            total_row = len(items or []) + 5
            ws.cell(row=total_row, column=5, value="TOTAL BALANCE").font = Font(bold=True)
            cell = ws.cell(row=total_row, column=6, value=total_balance)
            cell.font = Font(bold=True)
            cell.number_format = currency_fmt
            cell.alignment = right_align

        # ── Report: RPT Receivables (COA roll-forward statement) ──────────────
        elif data.report_type == "receivables":
            ws.title = "RPT Receivables"
            ws["A1"] = "MUNICIPAL TREASURY OFFICE — RPT RECEIVABLES STATEMENT"
            ws["A1"].font = title_font
            ws["A2"] = f"Report Year: {data.year}    Generated: {date_str}"
            ws["A2"].font = Font(italic=True, size=10)
            ws.merge_cells("A1:B1")
            ws.merge_cells("A2:B2")
            ws["A1"].alignment = center_align
            ws["A2"].alignment = center_align

            # Header row
            for col_idx, h in enumerate(["Line Item", "Amount"], 1):
                ws.cell(row=4, column=col_idx, value=h)
            style_header_row(ws, 4, 2)

            # Single computation path — same service the on-screen view uses.
            summary = bill_svc.get_rpt_receivables_summary(data.year, db_session=db_session)

            line_items = [
                ("Beginning Receivable",      float(summary.get("beginning_receivable", 0) or 0)),
                ("Current-Year Assessment",   float(summary.get("current_year_assessment", 0) or 0)),
                ("Collections",               float(summary.get("collections", 0) or 0)),
                ("Adjustments",               float(summary.get("adjustments", 0) or 0)),
                ("Ending Receivable",         float(summary.get("ending_receivable", 0) or 0)),
            ]

            for offset, (label, amount) in enumerate(line_items):
                row_idx = 5 + offset
                is_total = label == "Ending Receivable"
                label_cell = ws.cell(row=row_idx, column=1, value=label)
                label_cell.border = thin_border
                if is_total:
                    label_cell.font = Font(bold=True)
                amt_cell = ws.cell(row=row_idx, column=2, value=amount)
                amt_cell.number_format = currency_fmt
                amt_cell.alignment = right_align
                amt_cell.border = thin_border
                if is_total:
                    amt_cell.font = Font(bold=True)

        # ── Report: Receivables by Barangay ──────────────────────────────────
        elif data.report_type == "receivables_by_barangay":
            ws.title = "Receivables by Barangay"
            ws["A1"] = "MUNICIPAL TREASURY OFFICE — RECEIVABLES BY BARANGAY"
            ws["A1"].font = title_font
            
            brgy_filter = data.barangay if data.barangay and data.barangay != "ALL" else None
            brgy_lbl = f"Barangay: {brgy_filter}" if brgy_filter else "All Barangays"
            ws["A2"] = f"As of Year: {data.year}  ·  {brgy_lbl}    Generated: {date_str}"
            ws["A2"].font = Font(italic=True, size=10)
            ws.merge_cells("A1:G1")
            ws.merge_cells("A2:G2")
            ws["A1"].alignment = center_align
            ws["A2"].alignment = center_align

            headers = ["Barangay", "Assessed Value", "Total Due", "Penalty", "Discount", "Collected", "Total Receivable"]
            for col_idx, h in enumerate(headers, 1):
                ws.cell(row=4, column=col_idx, value=h)
            style_header_row(ws, 4, len(headers))

            y_val = None if data.year == "All" or data.year == "All Years" else int(data.year)
            rows = prop_svc.get_receivables_by_barangay(report_year=y_val, db_session=db_session)
            
            if brgy_filter:
                rows = [r for r in rows if r[0] == brgy_filter]

            grand_total = 0.0
            for row_idx, row in enumerate(rows or [], start=5):
                vals = [
                    row[0],
                    float(row[1] or 0),
                    float(row[2] or 0),
                    float(row[3] or 0),
                    float(row[4] or 0),
                    float(row[5] or 0),
                    float(row[6] or 0),
                ]
                for col_idx, val in enumerate(vals, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    cell.border = thin_border
                    if col_idx >= 2:
                        cell.number_format = currency_fmt
                        cell.alignment = right_align

                grand_total += float(row[6] or 0)

            total_row = len(rows or []) + 5
            ws.cell(row=total_row, column=1, value="GRAND TOTAL").font = Font(bold=True)
            cell = ws.cell(row=total_row, column=7, value=grand_total)
            cell.font = Font(bold=True)
            cell.number_format = currency_fmt
            cell.alignment = right_align

        # ── Report: Assessment Roll ───────────────────────────────────────────
        else:
            ws.title = "Assessment Roll"
            ws["A1"] = "MUNICIPAL TREASURY OFFICE — ASSESSMENT ROLL"
            ws["A1"].font = title_font
            
            brgy_filter = data.barangay if data.barangay and data.barangay != "ALL" else None
            brgy_lbl = f"Barangay: {brgy_filter}" if brgy_filter else "All Barangays"
            ws["A2"] = f"{brgy_lbl}    Generated: {date_str}"
            ws["A2"].font = Font(italic=True, size=10)
            ws.merge_cells("A1:F1")
            ws.merge_cells("A2:F2")
            ws["A1"].alignment = center_align
            ws["A2"].alignment = center_align

            headers = ["TD Number", "Owner Name", "Barangay", "Kind of Property",
                       "Assessed Value", "Tax Year"]
            for col_idx, h in enumerate(headers, 1):
                ws.cell(row=4, column=col_idx, value=h)
            style_header_row(ws, 4, len(headers))

            result = prop_svc.get_assessment_roll(limit=10000, cursor=None, db_session=db_session)
            items = result.get("items", []) if isinstance(result, dict) else result

            if brgy_filter:
                filtered_items = []
                for item in items:
                    if isinstance(item, dict):
                        b = item.get("barangay", "")
                    else:
                        b = item[6] if len(item) > 6 else ""
                    if b == brgy_filter:
                        filtered_items.append(item)
                items = filtered_items

            for row_idx, item in enumerate(items or [], start=5):
                if isinstance(item, dict):
                    vals = [
                        item.get("td_number", ""), item.get("owner_name", ""),
                        item.get("barangay", ""), item.get("kind_of_property", ""),
                        float(item.get("assessed_value", 0) or 0),
                        item.get("tax_year", ""),
                    ]
                else:
                    vals = list(item) + [""] * max(0, 6 - len(item))

                for col_idx, val in enumerate(vals, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    cell.border = thin_border
                    if col_idx == 5:
                        cell.number_format = currency_fmt
                        cell.alignment = right_align

        auto_width(ws)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.read()

    # Run the CPU-bound workbook build in a thread to avoid blocking the event loop
    excel_bytes = await asyncio.to_thread(_build_workbook)

    report_label = data.report_type.replace("_", "-")
    period_label = f"{data.month}-{data.year}" if data.report_type == "collections" else "all"
    filename = f"MTO_{report_label}_{period_label}_{datetime.now(timezone.utc).strftime('%Y%m%d')}.xlsx"

    mto_logger.info(
        f"Excel export generated: {filename}",
        user=current_user.get("username"),
        report_type=data.report_type,
    )

    return StreamingResponse(
        io.BytesIO(excel_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
