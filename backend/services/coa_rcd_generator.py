# -*- coding: utf-8 -*-
import io
import os
from datetime import datetime
from decimal import Decimal
from typing import Dict, Any, List
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session
from backend.models import Payment, BankDeposit, Property

def is_current_year(payment_tax_year: str, date_paid: datetime) -> bool:
    """
    Determines if the payment's tax year is Current or Prior.
    If the tax year is greater than or equal to the year of date_paid, it is Current (or Advance).
    Otherwise, it is Prior.
    """
    if not payment_tax_year or not date_paid:
        return True
    try:
        # Strip any formatting/ranges (e.g., '2026-2027') to get the base year
        year_str = str(payment_tax_year).split('-')[0].strip()
        tax_yr = int(year_str)
        return tax_yr >= date_paid.year
    except ValueError:
        return True

def generate_coa_rcd_excel(
    db: Session,
    start_date: datetime,
    end_date: datetime,
    liquidating_officer: str = "N/A",
    treasurer: str = "N/A"
) -> io.BytesIO:
    """
    Generates a print-ready, professional COA Appendix 39 Excel report
    containing three sheets:
      1. Summary (Form 39 Layout)
      2. Detailed Collections
      3. UACS Summary
    """
    # Fetch payments in date range
    # Ensure datetime boundary covers the whole day
    start_dt = datetime(start_date.year, start_date.month, start_date.day, 0, 0, 0)
    end_dt = datetime(end_date.year, end_date.month, end_date.day, 23, 59, 59, 999999)

    payments = db.query(Payment).join(Property).filter(
        Payment.date_paid >= start_dt,
        Payment.date_paid <= end_dt
    ).order_by(Payment.date_paid, Payment.or_number).all()

    # Fetch deposits in date range
    deposits = db.query(BankDeposit).filter(
        BankDeposit.date_deposited >= start_dt,
        BankDeposit.date_deposited <= end_dt
    ).order_by(BankDeposit.date_deposited).all()

    # Initialize calculations
    basic_current = Decimal("0.00")
    basic_prior = Decimal("0.00")
    basic_penalty = Decimal("0.00")
    basic_discount = Decimal("0.00")

    sef_current = Decimal("0.00")
    sef_prior = Decimal("0.00")
    sef_penalty = Decimal("0.00")
    sef_discount = Decimal("0.00")

    detailed_rows = []
    total_collections_amt = Decimal("0.00")

    for p in payments:
        amt = Decimal(str(p.amount or 0.00))
        pen = Decimal(str(p.penalty or 0.00))
        disc = Decimal(str(p.discount or 0.00))
        
        # 50/50 split between Basic and SEF
        b_amt = amt / 2
        s_amt = amt / 2
        
        b_pen = pen / 2
        s_pen = pen / 2
        
        b_disc = disc / 2
        s_disc = disc / 2
        
        net = amt + pen - disc
        total_collections_amt += net

        current = is_current_year(p.tax_year, p.date_paid)
        
        if current:
            basic_current += b_amt
            sef_current += s_amt
        else:
            basic_prior += b_amt
            sef_prior += s_amt
            
        basic_penalty += b_pen
        sef_penalty += s_pen
        
        basic_discount += b_disc
        sef_discount += s_disc

        payor_name = p.property.payor_name or p.property.owner_name or "N/A"
        date_str = p.date_paid.strftime("%Y-%m-%d") if p.date_paid else "N/A"
        
        detailed_rows.append({
            "date": date_str,
            "or_number": p.or_number or "N/A",
            "payor": payor_name,
            "basic": b_amt,
            "sef": s_amt,
            "penalty": pen,
            "discount": disc,
            "net": net
        })

    # Summary calculations for Section A (Collections)
    net_basic_current = basic_current - basic_discount
    net_sef_current = sef_current - sef_discount
    
    total_basic = basic_current + basic_prior + basic_penalty - basic_discount
    total_sef = sef_current + sef_prior + sef_penalty - sef_discount
    
    # Deposits calculation
    total_deposits_amt = Decimal("0.00")
    for d in deposits:
        total_deposits_amt += Decimal(str(d.amount or 0.00))

    # UACS code dictionary
    uacs_data = [
        {"code": "4-01-02-040-01", "desc": "Real Property Tax - Basic (Current Year)", "amount": basic_current},
        {"code": "4-01-02-040-02", "desc": "Real Property Tax - Basic (Prior Years)", "amount": basic_prior},
        {"code": "4-01-02-040-03", "desc": "Real Property Tax - Basic (Penalties)", "amount": basic_penalty},
        {"code": "4-01-02-040-04", "desc": "Discount on Real Property Tax - Basic", "amount": -basic_discount},
        {"code": "4-01-02-041-01", "desc": "Special Education Fund Tax (Current Year)", "amount": sef_current},
        {"code": "4-01-02-041-02", "desc": "Special Education Fund Tax (Prior Years)", "amount": sef_prior},
        {"code": "4-01-02-041-03", "desc": "Special Education Fund Tax (Penalties)", "amount": sef_penalty},
        {"code": "4-01-02-041-04", "desc": "Discount on Special Education Fund Tax", "amount": -sef_discount},
    ]

    # Create Workbook
    wb = Workbook()
    
    # ----------------------------------------------------
    # Styles Definition
    # ----------------------------------------------------
    font_family = "Segoe UI"
    
    font_title = Font(name=font_family, size=14, bold=True, color="1B365D")
    font_subtitle = Font(name=font_family, size=10, italic=True, color="555555")
    font_section = Font(name=font_family, size=11, bold=True, color="1B365D")
    font_header = Font(name=font_family, size=10, bold=True, color="FFFFFF")
    font_data = Font(name=font_family, size=10)
    font_total = Font(name=font_family, size=10, bold=True)
    font_bold = Font(name=font_family, size=10, bold=True)
    
    fill_header = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
    fill_section = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    fill_zebra = PatternFill(start_color="F2F4F7", end_color="F2F4F7", fill_type="solid")
    fill_accent = PatternFill(start_color="E9F0F8", end_color="E9F0F8", fill_type="solid")
    
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    
    border_thin = Side(border_style="thin", color="CCCCCC")
    border_double = Side(border_style="double", color="333333")
    border_thick_top = Side(border_style="medium", color="1B365D")
    
    border_cell = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)
    border_header = Border(left=border_thin, right=border_thin, top=border_thick_top, bottom=border_thick_top)
    border_total = Border(top=border_thin, bottom=border_double)

    # Format numbers: Philippine Peso symbol with 2 decimals
    currency_format = "₱#,##0.00;[Red](₱#,##0.00);\"₱0.00\""

    # ====================================================
    # SHEET 1: Summary (Form 39)
    # ====================================================
    ws1 = wb.active
    ws1.title = "Summary"
    ws1.views.sheetView[0].showGridLines = True

    # Municipality Title & Header
    mun_name = os.getenv("MUNICIPALITY_NAME", "DIPACULAO")
    ws1.cell(row=1, column=1, value="REPUBLIC OF THE PHILIPPINES").font = font_bold
    ws1.cell(row=2, column=1, value=f"MUNICIPALITY OF {mun_name.upper()}").font = font_bold
    ws1.cell(row=3, column=1, value="OFFICE OF THE MUNICIPAL TREASURER").font = font_bold
    
    ws1.cell(row=5, column=1, value="REPORT OF COLLECTIONS AND DEPOSITS").font = font_title
    ws1.cell(row=6, column=1, value=f"Period: {start_date.strftime('%B %d, %Y')} to {end_date.strftime('%B %d, %Y')}").font = font_subtitle

    # Section A: Collections Header
    ws1.row_dimensions[8].height = 24
    ws1.merge_cells("A8:C8")
    ws1.cell(row=8, column=1, value="SECTION A: COLLECTIONS SUMMARY").font = font_section
    ws1.cell(row=8, column=1).fill = fill_section
    ws1.cell(row=8, column=1).alignment = align_left

    # Subheadings for Section A
    headers_a = ["Account Classification / Source", "UACS Code", "Amount"]
    ws1.row_dimensions[9].height = 20
    for col_idx, h in enumerate(headers_a, 1):
        cell = ws1.cell(row=9, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center if col_idx == 2 else (align_right if col_idx == 3 else align_left)
        cell.border = border_header

    # Data rows for collections
    col_rows = [
        ("Real Property Tax - Basic (Current Year)", "4-01-02-040-01", basic_current),
        ("Real Property Tax - Basic (Prior Years)", "4-01-02-040-02", basic_prior),
        ("Real Property Tax - Basic (Penalties)", "4-01-02-040-03", basic_penalty),
        ("Discount on Real Property Tax - Basic", "4-01-02-040-04", -basic_discount),
        ("Special Education Fund Tax (Current Year)", "4-01-02-041-01", sef_current),
        ("Special Education Fund Tax (Prior Years)", "4-01-02-041-02", sef_prior),
        ("Special Education Fund Tax (Penalties)", "4-01-02-041-03", sef_penalty),
        ("Discount on Special Education Fund Tax", "4-01-02-041-04", -sef_discount),
    ]

    current_row = 10
    for title, uacs, val in col_rows:
        ws1.row_dimensions[current_row].height = 18
        
        c1 = ws1.cell(row=current_row, column=1, value=title)
        c2 = ws1.cell(row=current_row, column=2, value=uacs)
        c3 = ws1.cell(row=current_row, column=3, value=float(val))
        
        c1.font = font_data; c1.border = border_cell; c1.alignment = align_left
        c2.font = font_data; c2.border = border_cell; c2.alignment = align_center
        c3.font = font_data; c3.border = border_cell; c3.alignment = align_right
        c3.number_format = currency_format
        
        if current_row % 2 == 1:
            c1.fill = fill_zebra; c2.fill = fill_zebra; c3.fill = fill_zebra
        current_row += 1

    # Total Collections row
    ws1.row_dimensions[current_row].height = 20
    ws1.cell(row=current_row, column=1, value="TOTAL COLLECTIONS").font = font_total
    ws1.cell(row=current_row, column=1).alignment = align_left
    ws1.cell(row=current_row, column=1).border = border_total
    
    ws1.cell(row=current_row, column=2, value="").border = border_total
    
    total_cell = ws1.cell(row=current_row, column=3, value=float(total_collections_amt))
    total_cell.font = font_total
    total_cell.alignment = align_right
    total_cell.number_format = currency_format
    total_cell.border = border_total
    total_cell.fill = fill_accent

    # Section B: Deposits Header
    current_row += 3
    ws1.row_dimensions[current_row].height = 24
    ws1.merge_cells(f"A{current_row}:D{current_row}")
    sec_b_header = ws1.cell(row=current_row, column=1, value="SECTION B: BANK DEPOSITS SUMMARY")
    sec_b_header.font = font_section
    sec_b_header.fill = fill_section
    sec_b_header.alignment = align_left
    
    current_row += 1
    # Subheadings for Section B
    headers_b = ["Date Deposited", "Bank Name / Branch", "Deposit Slip / Reference", "Amount"]
    ws1.row_dimensions[current_row].height = 20
    for col_idx, h in enumerate(headers_b, 1):
        cell = ws1.cell(row=current_row, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center if col_idx in [1, 3] else (align_right if col_idx == 4 else align_left)
        cell.border = border_header

    start_dep_row = current_row + 1
    current_row += 1
    
    if not deposits:
        ws1.row_dimensions[current_row].height = 20
        ws1.merge_cells(f"A{current_row}:D{current_row}")
        no_dep_cell = ws1.cell(row=current_row, column=1, value="No bank deposits recorded for this period.")
        no_dep_cell.font = font_subtitle
        no_dep_cell.alignment = align_center
        for col_idx in range(1, 5):
            ws1.cell(row=current_row, column=col_idx).border = border_cell
        current_row += 1
    else:
        for idx, d in enumerate(deposits):
            ws1.row_dimensions[current_row].height = 18
            date_dep_str = d.date_deposited.strftime("%Y-%m-%d") if d.date_deposited else "N/A"
            
            c1 = ws1.cell(row=current_row, column=1, value=date_dep_str)
            c2 = ws1.cell(row=current_row, column=2, value=d.bank_name)
            c3 = ws1.cell(row=current_row, column=3, value=d.reference_number)
            c4 = ws1.cell(row=current_row, column=4, value=float(d.amount or 0.00))
            
            c1.font = font_data; c1.border = border_cell; c1.alignment = align_center
            c2.font = font_data; c2.border = border_cell; c2.alignment = align_left
            c3.font = font_data; c3.border = border_cell; c3.alignment = align_center
            c4.font = font_data; c4.border = border_cell; c4.alignment = align_right
            c4.number_format = currency_format
            
            if idx % 2 == 1:
                for col in [c1, c2, c3, c4]:
                    col.fill = fill_zebra
            current_row += 1

    # Total Deposits row
    ws1.row_dimensions[current_row].height = 20
    ws1.cell(row=current_row, column=1, value="TOTAL DEPOSITS").font = font_total
    ws1.cell(row=current_row, column=1).alignment = align_left
    ws1.cell(row=current_row, column=1).border = border_total
    
    for c_idx in [2, 3]:
        ws1.cell(row=current_row, column=c_idx).border = border_total
        
    dep_total_cell = ws1.cell(row=current_row, column=4, value=float(total_deposits_amt))
    dep_total_cell.font = font_total
    dep_total_cell.alignment = align_right
    dep_total_cell.number_format = currency_format
    dep_total_cell.border = border_total
    dep_total_cell.fill = fill_accent
    
    # Section C: Reconciliation
    current_row += 3
    ws1.row_dimensions[current_row].height = 24
    ws1.merge_cells(f"A{current_row}:C{current_row}")
    sec_c_header = ws1.cell(row=current_row, column=1, value="SECTION C: RECONCILIATION")
    sec_c_header.font = font_section
    sec_c_header.fill = fill_section
    sec_c_header.alignment = align_left
    
    current_row += 1
    ws1.row_dimensions[current_row].height = 18
    ws1.cell(row=current_row, column=1, value="Total Collections (Section A)").font = font_data
    c_coll = ws1.cell(row=current_row, column=3, value=float(total_collections_amt))
    c_coll.font = font_data; c_coll.number_format = currency_format; c_coll.alignment = align_right
    
    current_row += 1
    ws1.row_dimensions[current_row].height = 18
    ws1.cell(row=current_row, column=1, value="Less: Total Bank Deposits (Section B)").font = font_data
    c_dep = ws1.cell(row=current_row, column=3, value=float(total_deposits_amt))
    c_dep.font = font_data; c_dep.number_format = currency_format; c_dep.alignment = align_right
    
    current_row += 1
    ws1.row_dimensions[current_row].height = 20
    ws1.cell(row=current_row, column=1, value="UN-DEPOSITED BALANCE").font = font_total
    c_bal = ws1.cell(row=current_row, column=3, value=float(total_collections_amt - total_deposits_amt))
    c_bal.font = font_total; c_bal.number_format = currency_format; c_bal.alignment = align_right
    c_bal.border = Border(top=border_thin, bottom=border_double)
    if (total_collections_amt - total_deposits_amt) != 0:
        c_bal.fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid") # soft red highlight

    # Signatures
    current_row += 4
    ws1.row_dimensions[current_row].height = 18
    ws1.cell(row=current_row, column=1, value="Certified Correct:").font = font_bold
    ws1.cell(row=current_row, column=3, value="Acknowledged Receipt:").font = font_bold
    
    current_row += 3
    ws1.row_dimensions[current_row].height = 18
    ws1.cell(row=current_row, column=1, value=liquidating_officer.upper()).font = font_bold
    ws1.cell(row=current_row, column=3, value=treasurer.upper()).font = font_bold
    
    current_row += 1
    ws1.row_dimensions[current_row].height = 15
    ws1.cell(row=current_row, column=1, value="Liquidating Officer").font = font_subtitle
    ws1.cell(row=current_row, column=3, value="Municipal Treasurer").font = font_subtitle

    # Auto-adjust column widths for Summary
    for col in ws1.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            # Avoid using merged cells for column width calculation
            if cell.coordinate in ["A8", "B8", "C8", "A19", "B19", "C19", "D19", "A23", "B23", "C23"]:
                continue
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws1.column_dimensions[col_letter].width = max(max_len + 4, 15)
    ws1.column_dimensions["A"].width = 42
    ws1.column_dimensions["B"].width = 20
    ws1.column_dimensions["C"].width = 22
    ws1.column_dimensions["D"].width = 22

    # ====================================================
    # SHEET 2: Detailed Collections
    # ====================================================
    ws2 = wb.create_sheet(title="Detailed Collections")
    ws2.views.sheetView[0].showGridLines = True
    
    ws2.cell(row=1, column=1, value="DETAILED REPORT OF COLLECTIONS").font = font_title
    ws2.cell(row=2, column=1, value=f"Period: {start_date.strftime('%B %d, %Y')} to {end_date.strftime('%B %d, %Y')}").font = font_subtitle
    
    headers_detailed = ["Date Paid", "OR Number", "Payor Name", "Basic Tax (50%)", "SEF Tax (50%)", "Penalties", "Discounts", "Net Cash"]
    ws2.row_dimensions[4].height = 24
    for col_idx, h in enumerate(headers_detailed, 1):
        cell = ws2.cell(row=4, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center if col_idx in [1, 2] else (align_left if col_idx == 3 else align_right)
        cell.border = border_header

    row_idx = 5
    for idx, r in enumerate(detailed_rows):
        ws2.row_dimensions[row_idx].height = 18
        
        c1 = ws2.cell(row=row_idx, column=1, value=r["date"])
        c2 = ws2.cell(row=row_idx, column=2, value=r["or_number"])
        c3 = ws2.cell(row=row_idx, column=3, value=r["payor"])
        c4 = ws2.cell(row=row_idx, column=4, value=float(r["basic"]))
        c5 = ws2.cell(row=row_idx, column=5, value=float(r["sef"]))
        c6 = ws2.cell(row=row_idx, column=6, value=float(r["penalty"]))
        c7 = ws2.cell(row=row_idx, column=7, value=float(r["discount"]))
        c8 = ws2.cell(row=row_idx, column=8, value=float(r["net"]))
        
        c1.font = font_data; c1.alignment = align_center; c1.border = border_cell
        c2.font = font_data; c2.alignment = align_center; c2.border = border_cell
        c3.font = font_data; c3.alignment = align_left;   c3.border = border_cell
        
        for col_cell in [c4, c5, c6, c7, c8]:
            col_cell.font = font_data
            col_cell.alignment = align_right
            col_cell.border = border_cell
            col_cell.number_format = currency_format

        if idx % 2 == 1:
            for col_cell in [c1, c2, c3, c4, c5, c6, c7, c8]:
                col_cell.fill = fill_zebra
                
        row_idx += 1

    # Totals Row
    ws2.row_dimensions[row_idx].height = 20
    ws2.cell(row=row_idx, column=1, value="TOTAL").font = font_total
    ws2.cell(row=row_idx, column=1).alignment = align_left
    ws2.cell(row=row_idx, column=1).border = border_total
    
    for c_idx in [2, 3]:
        ws2.cell(row=row_idx, column=c_idx).border = border_total
        
    for col_idx, col_let in enumerate(["D", "E", "F", "G", "H"], 4):
        total_cell = ws2.cell(row=row_idx, column=col_idx, value=f"=SUM({col_let}5:{col_let}{row_idx-1})")
        total_cell.font = font_total
        total_cell.alignment = align_right
        total_cell.number_format = currency_format
        total_cell.border = border_total
        total_cell.fill = fill_accent

    # Auto-adjust column widths for Detailed tab
    for col in ws2.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.coordinate in ["A1", "A2", "B1", "B2"]:
                continue
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws2.column_dimensions[col_letter].width = max(max_len + 4, 14)
    ws2.column_dimensions["C"].width = 28

    # ====================================================
    # SHEET 3: UACS Summary
    # ====================================================
    ws3 = wb.create_sheet(title="UACS Summary")
    ws3.views.sheetView[0].showGridLines = True
    
    ws3.cell(row=1, column=1, value="RECONCILIATION BY UACS ACCOUNTS").font = font_title
    ws3.cell(row=2, column=1, value=f"Period: {start_date.strftime('%B %d, %Y')} to {end_date.strftime('%B %d, %Y')}").font = font_subtitle
    
    headers_uacs = ["UACS Account Code", "Account Description", "Amount (Debit/Credit)"]
    ws3.row_dimensions[4].height = 24
    for col_idx, h in enumerate(headers_uacs, 1):
        cell = ws3.cell(row=4, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center if col_idx == 1 else (align_left if col_idx == 2 else align_right)
        cell.border = border_header

    uacs_row = 5
    for idx, u in enumerate(uacs_data):
        ws3.row_dimensions[uacs_row].height = 18
        
        c1 = ws3.cell(row=uacs_row, column=1, value=u["code"])
        c2 = ws3.cell(row=uacs_row, column=2, value=u["desc"])
        c3 = ws3.cell(row=uacs_row, column=3, value=float(u["amount"]))
        
        c1.font = font_data; c1.alignment = align_center; c1.border = border_cell
        c2.font = font_data; c2.alignment = align_left;   c2.border = border_cell
        c3.font = font_data; c3.alignment = align_right;  c3.border = border_cell
        c3.number_format = currency_format
        
        if idx % 2 == 1:
            c1.fill = fill_zebra; c2.fill = fill_zebra; c3.fill = fill_zebra
            
        uacs_row += 1

    # Total row for UACS Summary
    ws3.row_dimensions[uacs_row].height = 20
    ws3.cell(row=uacs_row, column=1, value="TOTAL REVENUE").font = font_total
    ws3.cell(row=uacs_row, column=1).alignment = align_left
    ws3.cell(row=uacs_row, column=1).border = border_total
    
    ws3.cell(row=uacs_row, column=2).border = border_total
    
    total_uacs_cell = ws3.cell(row=uacs_row, column=3, value=f"=SUM(C5:C{uacs_row-1})")
    total_uacs_cell.font = font_total
    total_uacs_cell.alignment = align_right
    total_uacs_cell.number_format = currency_format
    total_uacs_cell.border = border_total
    total_uacs_cell.fill = fill_accent

    # Auto-adjust column widths for UACS tab
    for col in ws3.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.coordinate in ["A1", "A2", "B1", "B2"]:
                continue
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws3.column_dimensions[col_letter].width = max(max_len + 4, 15)
    ws3.column_dimensions["A"].width = 24
    ws3.column_dimensions["B"].width = 48
    ws3.column_dimensions["C"].width = 24

    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
