import customtkinter as ctk
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime
import threading
import shutil
import os
from theme_manager import ModernTheme
import api_clients.billing_service as billing
import api_clients.property_service as prop
import api_clients.system_service as system
import api_clients.reports_service as reports_api
from ui_components import ErrorDialog, show_toast
from utils import format_curr, tr


class ReportsPage:
    def __init__(self, parent, user):
        self.parent = parent
        self.user = user
        self._last_reconciliation_payload = None
        self.setup_ui()

    def setup_ui(self):
        # Progress bar at the very top (outside main container)
        self.loading_bar = ctk.CTkProgressBar(self.parent, height=2, corner_radius=0, progress_color=ModernTheme.PRIMARY)
        self.loading_bar.pack(fill="x")
        self.loading_bar.set(0)
        self.loading_bar.pack_forget()

        self.container = ctk.CTkFrame(self.parent, fg_color="transparent")
        self.container.pack(fill="both", expand=True, padx=20, pady=20)

        # Header
        ctk.CTkLabel(
            self.container,
            text=tr("reports.title"),
            font=ModernTheme.H2,
        ).pack(anchor="w", pady=(0, 20))

        # A stable, equal-width report navigator is easier to scan than the
        # compact default segmented control.
        self.tabview = ctk.CTkTabview(
            self.container,
            anchor="nw",
            corner_radius=0,
            border_width=0,
            fg_color="transparent",
            segmented_button_fg_color="#334155",
            segmented_button_selected_color=ModernTheme.PRIMARY,
            segmented_button_selected_hover_color="#0284c7",
            segmented_button_unselected_color="#334155",
            segmented_button_unselected_hover_color="#475569",
            text_color="#f8fafc",
        )
        self.tabview.pack(fill="both", expand=True)

        self.collection_tab = self.tabview.add(tr("reports.tabs.collection"))
        self.receivables_tab = self.tabview.add(tr("reports.tabs.receivables"))
        self.barangay_tab = self.tabview.add(tr("reports.tabs.barangay"))
        self.reconciliation_tab = self.tabview.add(tr("reports.tabs.reconciliation"))

        # Keep all report destinations visually balanced and comfortably clickable.
        self.tabview._segmented_button.configure(
            width=1040,
            height=42,
            corner_radius=6,
            border_width=1,
            font=("Inter", 12, "bold"),
            dynamic_resizing=False,
        )

        self.setup_collection_tab()
        self.setup_receivables_tab()
        self.setup_barangay_tab()
        self.setup_reconciliation_tab()

    def _show_loading(self):
        try:
            self.loading_bar.pack(fill="x", side="top", before=self.container)
            self.loading_bar.start()
        except Exception:
            pass

    def _hide_loading(self):
        try:
            self.loading_bar.stop()
            self.loading_bar.pack_forget()
        except Exception:
            pass

    def _bind_enter(self, widget, callback):
        def submit(_event=None):
            callback()
            return "break"

        try:
            widget.bind("<Return>", submit)
            widget.bind("<KP_Enter>", submit)
        except Exception:
            pass

        inner_entry = getattr(widget, "_entry", None)
        if inner_entry is not None:
            try:
                inner_entry.bind("<Return>", submit)
                inner_entry.bind("<KP_Enter>", submit)
            except Exception:
                pass

    def setup_collection_tab(self):
        filter_fr = ctk.CTkFrame(self.collection_tab, fg_color=ModernTheme.SECONDARY, corner_radius=8)
        filter_fr.pack(fill="x", padx=10, pady=10)

        ctk.CTkLabel(
            filter_fr, text=tr("reports.collection.filter_label"), font=ModernTheme.BODY_BOLD, text_color="white"
        ).pack(side="left", padx=10)

        self.month_cb = ctk.CTkComboBox(
            filter_fr, values=["All"] + [f"{m:02d}" for m in range(1, 13)]
        )
        self.month_cb.set(datetime.now().strftime("%m"))
        self.month_cb.pack(side="left", padx=5)

        current_year = datetime.now().year
        self.year_cb = ctk.CTkComboBox(
            filter_fr,
            values=["All"]
            + [str(y) for y in range(current_year - 10, current_year + 3)],
        )
        self.year_cb.set(str(current_year))
        self.year_cb.pack(side="left", padx=5)
        self._bind_enter(self.month_cb, self.generate_collection_report)
        self._bind_enter(self.year_cb, self.generate_collection_report)

        ctk.CTkButton(
            filter_fr, text=tr("reports.collection.btn_generate"), command=self.generate_collection_report,
            font=ModernTheme.BUTTON, fg_color=ModernTheme.SUCCESS
        ).pack(side="left", padx=10)

        # 🏦 Manage Bank Deposits button
        self.manage_dep_btn = ctk.CTkButton(
            filter_fr, text="🏦 Manage Deposits", command=self.open_manage_deposits,
            font=ModernTheme.BUTTON, fg_color=ModernTheme.PRIMARY
        )
        self.manage_dep_btn.pack(side="left", padx=10)

        # 📊 Export COA RCD button
        self.export_rcd_btn = ctk.CTkButton(
            filter_fr, text="📊 Export COA RCD (Excel)", command=self.open_export_signatories,
            font=ModernTheme.BUTTON, fg_color=ModernTheme.WARNING
        )
        role = str(self.user.get("role", "")).strip().lower() if isinstance(self.user, dict) else ""
        if role == "admin":
            self.export_rcd_btn.pack(side="left", padx=10)

        # ── Pagination state ──────────────────────────────────────────────────
        self._coll_page_size = 100
        self._coll_cursors = [None]   # index 0 = first page cursor (None)
        self._coll_page = 0
        self._coll_has_more = False

        self.coll_table_fr = ctk.CTkFrame(self.collection_tab)
        self.coll_table_fr.pack(fill="both", expand=True, padx=10, pady=(10, 0))

        style = ttk.Style()
        style.theme_use("clam")
        style.configure(
            "Treeview",
            rowheight=35,
            font=ModernTheme.BODY,
            background="#1e1e1e",
            fieldbackground="#1e1e1e",
            foreground="white",
        )
        style.configure(
            "Treeview.Heading",
            font=ModernTheme.BODY_BOLD,
            background="#333333",
            foreground="white",
        )

        cols = (
            tr("reports.collection.table.date"),
            tr("reports.collection.table.or"),
            tr("reports.collection.table.td"),
            tr("reports.collection.table.owner"),
            tr("reports.collection.table.kind"),
            tr("reports.collection.table.year"),
            tr("reports.collection.table.amount"),
            tr("reports.collection.table.posted")
        )

        tree_container = tk.Frame(self.coll_table_fr, bg="#1e1e1e")
        tree_container.pack(fill="both", expand=True)

        scrolly = ttk.Scrollbar(tree_container, orient="vertical")
        scrollx = ttk.Scrollbar(tree_container, orient="horizontal")
        scrolly.pack(side="right", fill="y")
        scrollx.pack(side="bottom", fill="x")

        self.coll_tree = ttk.Treeview(
            tree_container, columns=cols, show="headings",
            yscrollcommand=scrolly.set, xscrollcommand=scrollx.set,
        )
        scrolly.configure(command=self.coll_tree.yview)
        scrollx.configure(command=self.coll_tree.xview)

        for col in cols:
            self.coll_tree.heading(col, text=col.upper())
            self.coll_tree.column(col, width=110, anchor="center")
        self.coll_tree.column(tr("reports.collection.table.owner"), width=180, anchor="w")
        self.coll_tree.pack(side="left", fill="both", expand=True)

        self.coll_tree.tag_configure("oddrow",  background="#2b2b2b", foreground="white")
        self.coll_tree.tag_configure("evenrow", background="#333333", foreground="white")

        # ── Pagination bar ────────────────────────────────────────────────────
        pag_fr = ctk.CTkFrame(self.collection_tab, fg_color="transparent")
        pag_fr.pack(fill="x", padx=10, pady=(4, 10))

        self._coll_prev_btn = ctk.CTkButton(
            pag_fr, text="◀  PREVIOUS",
            command=self._coll_prev_page,
            width=120, height=32,
            font=ModernTheme.BUTTON_SMALL,
            fg_color=ModernTheme.SECONDARY,
            hover_color=ModernTheme.SECONDARY_HOVER,
            state="disabled",
        )
        self._coll_prev_btn.pack(side="left", padx=(0, 8))

        self._coll_page_lbl = ctk.CTkLabel(
            pag_fr, text="Page 1",
            font=("Inter", 11, "bold"),
            text_color=ModernTheme.TEXT_GRAY,
        )
        self._coll_page_lbl.pack(side="left", expand=True)

        self._coll_next_btn = ctk.CTkButton(
            pag_fr, text="NEXT  ▶",
            command=self._coll_next_page,
            width=120, height=32,
            font=ModernTheme.BUTTON_SMALL,
            fg_color=ModernTheme.SECONDARY,
            hover_color=ModernTheme.SECONDARY_HOVER,
            state="disabled",
        )
        self._coll_next_btn.pack(side="right")

    def setup_receivables_tab(self):
        receiv_fr = ctk.CTkFrame(self.receivables_tab, fg_color="transparent")
        receiv_fr.pack(fill="both", expand=True, padx=10, pady=10)

        filter_fr = ctk.CTkFrame(receiv_fr, fg_color="transparent")
        filter_fr.pack(fill="x", pady=(0, 12))

        curr_y = datetime.now().year
        self.receiv_year_cb = ctk.CTkComboBox(
            filter_fr, values=[str(y) for y in range(curr_y - 10, curr_y + 3)]
        )
        self.receiv_year_cb.set(str(curr_y))
        self.receiv_year_cb.pack(side="left", padx=10)
        self._bind_enter(self.receiv_year_cb, self.generate_receivables_report)

        ctk.CTkButton(
            filter_fr, text=tr("reports.receivables.btn_load"), command=self.generate_receivables_report,
            font=ModernTheme.BUTTON, fg_color=ModernTheme.PRIMARY, height=34
        ).pack(side="left", padx=(8, 0))

        self.receiv_export_excel_btn = ctk.CTkButton(
            filter_fr,
            text="EXPORT EXCEL",
            command=self._export_receivables_excel,
            font=ModernTheme.BUTTON,
            fg_color=ModernTheme.SUCCESS,
            height=34,
            width=130,
        )
        self.receiv_export_excel_btn.pack(side="right", padx=(8, 0))

        self.receiv_export_brgy_btn = ctk.CTkButton(
            filter_fr,
            text="BARANGAY PDF",
            command=self._export_receivables_barangay_pdf,
            font=ModernTheme.BUTTON,
            fg_color=ModernTheme.SECONDARY,
            height=34,
            width=130,
        )
        self.receiv_export_brgy_btn.pack(side="right")

        self.receiv_content = ctk.CTkFrame(receiv_fr, fg_color="transparent")
        self.receiv_content.pack(fill="both", expand=True)
        self.receiv_label = ctk.CTkLabel(
            self.receiv_content, text=tr("reports.receivables.hint"), font=ModernTheme.BODY, text_color=ModernTheme.TEXT_GRAY
        )
        self.receiv_label.pack(pady=50)

    def setup_barangay_tab(self):
        brgy_fr = ctk.CTkFrame(self.barangay_tab, fg_color="transparent")
        brgy_fr.pack(fill="both", expand=True, padx=20, pady=20)

        # ── Header with year filter ──────────────────────────────────────────
        top_fr = ctk.CTkFrame(brgy_fr, fg_color="transparent")
        top_fr.pack(fill="x", pady=(0, 8))

        ctk.CTkLabel(
            top_fr,
            text=tr("reports.barangay.title"),
            font=ModernTheme.H3,
            text_color=ModernTheme.PRIMARY,
        ).pack(side="left")

        # Year filter + export buttons — right side
        filter_fr = ctk.CTkFrame(top_fr, fg_color="transparent")
        filter_fr.pack(side="right")

        ctk.CTkLabel(
            filter_fr, text="As of Year:", font=ModernTheme.BODY, text_color="gray"
        ).pack(side="left", padx=(0, 8))

        curr_y = datetime.now().year
        self.brgy_year_cb = ctk.CTkComboBox(
            filter_fr,
            values=["All Years"] + [str(y) for y in range(curr_y - 10, curr_y + 1)],
            width=130,
            command=lambda _: self.generate_barangay_receivables(),
        )
        self.brgy_year_cb.set(str(curr_y))
        self.brgy_year_cb.pack(side="left", padx=(0, 8))
        self._bind_enter(self.brgy_year_cb, self.generate_barangay_receivables)



        ctk.CTkButton(
            filter_fr,
            text=f"🔄 {tr('reports.barangay.btn_refresh')}",
            command=self.generate_barangay_receivables,
            width=160,
            height=35,
            font=ModernTheme.BUTTON,
            fg_color=ModernTheme.SECONDARY,
        ).pack(side="left", padx=(0, 6))

        # ── Export buttons ───────────────────────────────────────────────────
        ctk.CTkButton(
            filter_fr,
            text="📄 Export PDF",
            command=self._export_brgy_pdf,
            width=130,
            height=35,
            font=ModernTheme.BUTTON,
            fg_color="#c0392b",
            hover_color="#962d22",
        ).pack(side="left", padx=(0, 6))

        ctk.CTkButton(
            filter_fr,
            text="📊 Export Excel",
            command=self._export_brgy_excel,
            width=140,
            height=35,
            font=ModernTheme.BUTTON,
            fg_color="#1a7431",
            hover_color="#145a27",
        ).pack(side="left")

        # Info banner explaining cumulative logic
        info_fr = ctk.CTkFrame(brgy_fr, fg_color=("#1a2634", "#1a2634"), corner_radius=8)
        info_fr.pack(fill="x", pady=(0, 10))
        ctk.CTkLabel(
            info_fr,
            text=(
                "ℹ️  Cumulative receivables: includes all unpaid balances from prior years "
                "carried forward into the selected year."
            ),
            font=("Segoe UI", 10),
            text_color="#8b949e",
            anchor="w",
        ).pack(padx=14, pady=8, anchor="w")

        # ── Table ────────────────────────────────────────────────────────────
        t_container = ctk.CTkFrame(brgy_fr)
        t_container.pack(fill="both", expand=True)

        cols = (
            tr("reports.barangay.table.barangay"),
            tr("reports.barangay.table.assessed"),
            tr("reports.barangay.table.due"),
            tr("reports.barangay.table.penalty"),
            tr("reports.barangay.table.discount"),
            tr("reports.barangay.table.collected"),
            tr("reports.barangay.table.receivable"),
        )
        self.brgy_tree = ttk.Treeview(t_container, columns=cols, show="headings")
        for col in cols:
            self.brgy_tree.heading(col, text=col.upper())
            self.brgy_tree.column(col, width=120, anchor="center")

        self.brgy_tree.column("Barangay", width=180, anchor="w")
        self.brgy_tree.column("Total Receivable", width=150)

        scrolly = ttk.Scrollbar(t_container, orient="vertical", command=self.brgy_tree.yview)
        self.brgy_tree.configure(yscrollcommand=scrolly.set)
        self.brgy_tree.pack(side="left", fill="both", expand=True)
        scrolly.pack(side="right", fill="y")

        self.brgy_tree.tag_configure("oddrow", background="#2b2b2b", foreground="white")
        self.brgy_tree.tag_configure("evenrow", background="#333333", foreground="white")

        # ── Summary footer ───────────────────────────────────────────────────
        self.brgy_summary = ctk.CTkFrame(brgy_fr, height=50, fg_color="#2c3e50", corner_radius=8)
        self.brgy_summary.pack(fill="x", pady=(15, 0))

        self.brgy_year_lbl = ctk.CTkLabel(
            self.brgy_summary,
            text="",
            font=("Segoe UI", 10),
            text_color="#8b949e",
        )
        self.brgy_year_lbl.pack(side="left", padx=20, pady=10)

        self.brgy_total_lbl = ctk.CTkLabel(
            self.brgy_summary,
            text=tr("reports.barangay.total").replace("{value}", "P 0.00"),
            font=ModernTheme.H3,
            text_color="white",
        )
        self.brgy_total_lbl.pack(side="right", padx=30, pady=10)

    def setup_reconciliation_tab(self):
        rec_fr = ctk.CTkFrame(self.reconciliation_tab, fg_color="transparent")
        rec_fr.pack(fill="both", expand=True, padx=10, pady=10)

        top_fr = ctk.CTkFrame(rec_fr, fg_color="transparent")
        top_fr.pack(fill="x", pady=(0, 12))

        title_fr = ctk.CTkFrame(top_fr, fg_color="transparent")
        title_fr.pack(side="left", fill="x", expand=True)
        ctk.CTkLabel(
            title_fr,
            text="INTER-DEPARTMENT RECONCILIATION",
            font=ModernTheme.H2,
            text_color=ModernTheme.PRIMARY,
        ).pack(anchor="w")
        ctk.CTkLabel(
            title_fr,
            text="COA-ready check: assessed levy should equal collections plus remaining receivables.",
            font=("Inter", 11),
            text_color=ModernTheme.TEXT_GRAY,
        ).pack(anchor="w", pady=(2, 0))

        controls = ctk.CTkFrame(top_fr, fg_color="transparent")
        controls.pack(side="right")
        curr_y = datetime.now().year
        self.recon_year_cb = ctk.CTkComboBox(
            controls,
            values=[str(y) for y in range(curr_y - 10, curr_y + 3)],
            width=120,
        )
        self.recon_year_cb.set(str(curr_y))
        self.recon_year_cb.pack(side="left", padx=(0, 8))
        self._bind_enter(self.recon_year_cb, self.generate_reconciliation_report)
        ctk.CTkButton(
            controls,
            text="LOAD CHECK",
            command=self.generate_reconciliation_report,
            font=ModernTheme.BUTTON,
            fg_color=ModernTheme.PRIMARY,
            height=34,
            width=130,
        ).pack(side="left")
        self.recon_export_excel_btn = ctk.CTkButton(
            controls,
            text="EXPORT EXCEL",
            command=self._export_reconciliation_excel,
            font=ModernTheme.BUTTON,
            fg_color=ModernTheme.SUCCESS,
            height=34,
            width=130,
            state="disabled",
        )
        self.recon_export_excel_btn.pack(side="left", padx=(8, 0))
        self.recon_export_pdf_btn = ctk.CTkButton(
            controls,
            text="PRINT PDF",
            command=self._export_reconciliation_pdf,
            font=ModernTheme.BUTTON,
            fg_color=ModernTheme.DANGER,
            height=34,
            width=110,
            state="disabled",
        )
        self.recon_export_pdf_btn.pack(side="left", padx=(8, 0))

        self.recon_content = ctk.CTkScrollableFrame(
            rec_fr,
            fg_color="transparent",
            scrollbar_button_color="#334155",
            scrollbar_button_hover_color="#475569",
        )
        self.recon_content.pack(fill="both", expand=True)
        ctk.CTkLabel(
            self.recon_content,
            text="Select a fiscal year, then load the reconciliation check.",
            font=ModernTheme.BODY,
            text_color=ModernTheme.TEXT_GRAY,
        ).pack(pady=60)

    def generate_reconciliation_report(self):
        year = self.recon_year_cb.get()
        self._show_loading()

        def worker():
            try:
                summary = billing.get_rpt_receivables_summary(year)
                metrics = billing.get_reconciliation_metrics(year)
                diagnostics = billing.get_reconciliation_diagnostics(year, limit=25)
                brgy_rows = prop.get_receivables_by_barangay(year=int(year))
                self.container.after(0, lambda: self._update_reconciliation(summary, brgy_rows, metrics, diagnostics))
            except Exception as e:
                self.container.after(0, lambda err=e: messagebox.showerror("Reconciliation Error", str(err)))
            finally:
                self.container.after(0, self._hide_loading)

        threading.Thread(target=worker, daemon=True).start()

    def _update_reconciliation(self, data, brgy_rows=None, metrics=None, diagnostics=None):
        for child in self.recon_content.winfo_children():
            child.destroy()

        if not data:
            self._last_reconciliation_payload = None
            self.recon_export_excel_btn.configure(state="disabled")
            self.recon_export_pdf_btn.configure(state="disabled")
            ctk.CTkLabel(
                self.recon_content,
                text="No reconciliation data available for the selected year.",
                font=ModernTheme.BODY,
                text_color=ModernTheme.TEXT_GRAY,
            ).pack(pady=60)
            return

        year = data.get("report_year", self.recon_year_cb.get())
        beginning = float(data.get("beginning_receivable", 0) or 0)
        current_net = float(data.get("current_year_net_collectible", data.get("current_year_assessment", 0)) or 0)
        current_penalty = float(data.get("current_year_penalty", 0) or 0)
        current_discount = float(data.get("current_year_discount", 0) or 0)
        adjustments = float(data.get("adjustments", 0) or 0)
        collections = float(data.get("collections", 0) or 0)
        calendar_applicable_collections = float(data.get("calendar_applicable_collections", collections) or 0)
        prepaid_current_year = float(data.get("prepaid_current_year", 0) or 0)
        future_year_prepayments = float(data.get("future_year_prepayments", 0) or 0)
        unpaid = float(data.get("ending_receivable", 0) or 0)
        levy = beginning + current_net + adjustments
        brgy_rows = brgy_rows or []
        metrics = metrics or {}
        diagnostics = diagnostics or {}
        self._last_reconciliation_payload = self._build_reconciliation_payload(
            data, brgy_rows, metrics, diagnostics
        )
        self.recon_export_excel_btn.configure(state="normal")
        self.recon_export_pdf_btn.configure(state="normal")
        assessor = metrics.get("assessor", {})
        treasury = metrics.get("treasury", {})
        delinquency = metrics.get("delinquency", {})

        # The reconciliation equation must use the same basis on all sides:
        # beginning receivable + net current collectible + adjustments = collections + ending receivable.
        current_levy = float(assessor.get("current_year_levy", data.get("current_year_levy", current_net)) or 0)
        current_penalty = float(assessor.get("current_year_penalty", current_penalty) or 0)
        current_discount = float(assessor.get("current_year_discount", current_discount) or 0)
        current_net = float(assessor.get("current_year_net_collectible", current_net) or 0)
        levy = beginning + current_net + adjustments
        total_collectible = levy
        cash_collected_this_year = float(treasury.get("cash_collected_this_year", collections) or 0)
        calendar_applicable_collections = float(treasury.get("calendar_applicable_collections", calendar_applicable_collections) or 0)
        prepaid_current_year = float(treasury.get("prepaid_current_year", prepaid_current_year) or 0)
        future_year_prepayments = float(treasury.get("future_year_prepayments", future_year_prepayments) or 0)
        treasury_total = float(treasury.get("total_collected", data.get("applied_collections", calendar_applicable_collections + prepaid_current_year)) or 0)
        expected_unpaid = float(data.get("expected_ending_receivable", unpaid) or 0)
        tracker_total = float(delinquency.get("total_unpaid", unpaid) or 0)
        current_receivable = float(delinquency.get("current_year_receivables", 0) or 0)
        prior_receivable = float(delinquency.get("prior_year_receivables", max(tracker_total - current_receivable, 0)) or 0)
        delinquency_total = float(delinquency.get("total_unpaid", prior_receivable + current_receivable) or 0)
        equation_variance = total_collectible - (treasury_total + delinquency_total)
        tracker_variance = float(diagnostics.get("tracker_variance", tracker_total - expected_unpaid) or 0)
        raw_tracker_variance = float(diagnostics.get("raw_tracker_variance", tracker_variance) or 0)
        variance = equation_variance if abs(equation_variance) > 1.0 else tracker_variance
        abs_variance = max(abs(equation_variance), abs(tracker_variance))
        tolerance = 1.0
        balanced = abs_variance <= tolerance
        collection_rate = (treasury_total / total_collectible * 100) if total_collectible > 0 else 0
        delinquency_rate = (expected_unpaid / total_collectible * 100) if total_collectible > 0 else 0

        def money(value):
            return f"P {float(value or 0):,.2f}"

        def stat_card(parent, col, label, value, color, subtitle):
            card = ctk.CTkFrame(
                parent,
                fg_color=("#e2e8f0", "#111827"),
                corner_radius=8,
                border_width=1,
                border_color=("#cbd5e1", "#243244"),
            )
            card.grid(row=0, column=col, sticky="nsew", padx=(0 if col == 0 else 8, 0 if col == 2 else 8))
            ctk.CTkLabel(card, text=label.upper(), font=("Inter", 10, "bold"), text_color=ModernTheme.TEXT_GRAY).pack(anchor="w", padx=16, pady=(14, 4))
            ctk.CTkLabel(card, text=value, font=("Inter", 22, "bold"), text_color=color).pack(anchor="w", padx=16)
            ctk.CTkLabel(card, text=subtitle, font=("Inter", 10), text_color=ModernTheme.TEXT_GRAY, wraplength=310, justify="left").pack(anchor="w", padx=16, pady=(4, 14))
            return card

        header = ctk.CTkFrame(self.recon_content, fg_color="transparent")
        header.pack(fill="x", pady=(0, 12))
        ctk.CTkLabel(
            header,
            text=f"FISCAL YEAR {year} RECONCILIATION CHECK",
            font=ModernTheme.H2,
            text_color=ModernTheme.PRIMARY,
        ).pack(side="left")
        status_color = ModernTheme.SUCCESS if balanced else ModernTheme.DANGER
        status_text = "BALANCED" if balanced else "NEEDS REVIEW"
        ctk.CTkLabel(header, text=status_text, font=ModernTheme.H3, text_color=status_color).pack(side="right")

        def field_row(parent, label, value, color="#f8fafc"):
            row = ctk.CTkFrame(parent, fg_color="transparent")
            row.pack(fill="x", padx=16, pady=(0, 10))
            ctk.CTkLabel(row, text=label, font=("Inter", 10, "bold"), text_color="#bfdbfe", anchor="w").pack(anchor="w")
            box = ctk.CTkFrame(row, fg_color="#0f172a", corner_radius=7, border_width=1, border_color="#334155")
            box.pack(fill="x", pady=(4, 0))
            ctk.CTkLabel(box, text=value, font=("Consolas", 13, "bold"), text_color=color, anchor="w").pack(fill="x", padx=12, pady=8)

        def department_card(parent, col, badge, title, accent, rows, footer_label, footer_value):
            card = ctk.CTkFrame(parent, fg_color=("#e2e8f0", "#111827"), corner_radius=8, border_width=1, border_color=("#cbd5e1", "#243244"))
            card.grid(row=0, column=col, sticky="nsew", padx=(0 if col == 0 else 8, 0 if col == 2 else 8))
            head = ctk.CTkFrame(card, fg_color="transparent")
            head.pack(fill="x", padx=16, pady=(14, 8))
            ctk.CTkLabel(head, text=badge, font=("Inter", 10, "bold"), text_color=accent, fg_color="#0f2747", corner_radius=5).pack(anchor="w")
            ctk.CTkLabel(card, text=title, font=("Inter", 16, "bold"), text_color="#f8fafc").pack(anchor="w", padx=16, pady=(0, 12))
            for label, value, color in rows:
                field_row(card, label, value, color)
            ctk.CTkFrame(card, height=1, fg_color="#243244").pack(fill="x", pady=(8, 0))
            footer = ctk.CTkFrame(card, fg_color="transparent")
            footer.pack(fill="x", padx=16, pady=14)
            ctk.CTkLabel(footer, text=footer_label, font=("Inter", 11), text_color="#bfdbfe").pack(side="left")
            ctk.CTkLabel(footer, text=footer_value, font=("Consolas", 13, "bold"), text_color=accent).pack(side="right")

        department_grid = ctk.CTkFrame(self.recon_content, fg_color="transparent")
        department_grid.pack(fill="x", pady=(0, 14))
        department_grid.grid_columnconfigure((0, 1, 2), weight=1)
        department_card(
            department_grid,
            0,
            "Assessor's office",
            "Tax levy data",
            ModernTheme.PRIMARY,
            (
                ("Total assessed value", money(assessor.get("total_assessed_value", 0)), "#f8fafc"),
                ("Basic tax rate", f"{float(assessor.get('tax_rate_percent', 0) or 0):.2f}%", "#f8fafc"),
                ("Total RPT rate", f"{float(assessor.get('total_tax_rate_percent', 0) or 0):.2f}%", "#f8fafc"),
                ("No. of taxable properties", f"{int(assessor.get('taxable_properties', 0) or 0):,}", "#f8fafc"),
                ("Penalties / interest", money(current_penalty), "#f59e0b"),
                ("Discounts", f"- {money(current_discount)}", "#22c55e"),
            ),
            "Current year levy",
            money(current_levy),
        )
        department_card(
            department_grid,
            1,
            "Treasurer's office",
            "Collection data",
            ModernTheme.SUCCESS,
            (
                ("Basic tax collected / credited", money(treasury.get("basic_tax_collected", 0)), ModernTheme.SUCCESS),
                ("Cash collected this fiscal year", money(cash_collected_this_year), "#f8fafc"),
                ("Current-year tax prepaid before year", money(prepaid_current_year), "#38bdf8"),
                ("Future prepayments excluded", money(future_year_prepayments), "#f59e0b"),
                ("No. of accounts paid", f"{int(treasury.get('accounts_paid', 0) or 0):,}", "#f8fafc"),
                ("No. of partial payments", f"{int(treasury.get('partial_payments', 0) or 0):,}", "#f8fafc"),
            ),
            "Applied collections / credits",
            money(treasury_total),
        )
        department_card(
            department_grid,
            2,
            "RPT tracker",
            "Delinquency data",
            ModernTheme.DANGER,
            (
                ("Prior-year receivables", money(prior_receivable), ModernTheme.DANGER),
                ("Current year receivables", money(current_receivable), "#f59e0b"),
                ("No. of delinquent accounts", f"{int(delinquency.get('delinquent_accounts', 0) or 0):,}", "#f8fafc"),
                ("Penalties and interest", money(delinquency.get("penalties_interest", 0)), "#f8fafc"),
            ),
            "Ending receivable",
            money(delinquency_total),
        )
        grid = ctk.CTkFrame(self.recon_content, fg_color="transparent")
        grid.pack(fill="x", pady=(0, 14))
        grid.grid_columnconfigure((0, 1, 2), weight=1)
        stat_card(grid, 0, "Total Collectible (A)", money(total_collectible), ModernTheme.PRIMARY, f"Beginning {money(beginning)} + net current {money(current_net)} + adjustments {money(adjustments)}")
        stat_card(grid, 1, "Collections / Credits (B)", money(treasury_total), ModernTheme.SUCCESS, f"Cash applied {money(calendar_applicable_collections)} + prepayments {money(prepaid_current_year)}")
        stat_card(grid, 2, "Ending Receivable (C)", money(delinquency_total), ModernTheme.DANGER, f"Receivable rate: {delinquency_rate:.1f}%")

        equation = ctk.CTkFrame(
            self.recon_content,
            fg_color=("#dbeafe", "#070b1a"),
            corner_radius=8,
            border_width=1,
            border_color=("#93c5fd", "#1e3a8a"),
        )
        equation.pack(fill="x", pady=(0, 14), padx=0)

        equation_top = ctk.CTkFrame(equation, fg_color="transparent")
        equation_top.pack(fill="x", padx=16, pady=(12, 4))
        ctk.CTkLabel(
            equation_top,
            text="RECONCILIATION EQUATION",
            font=("Inter", 10, "bold"),
            text_color="#93a4c7",
        ).pack(side="left")
        ctk.CTkLabel(
            equation_top,
            text="Total collectible should equal collections plus ending receivable",
            font=("Inter", 10),
            text_color="#bfdbfe",
        ).pack(side="right")

        formula = ctk.CTkFrame(equation, fg_color="transparent")
        formula.pack(fill="x", padx=16, pady=(0, 12))
        ctk.CTkLabel(
            formula,
            text=f"{money(total_collectible)}",
            font=("Consolas", 16, "bold"),
            text_color=ModernTheme.PRIMARY,
        ).pack(side="left")
        ctk.CTkLabel(formula, text="  =  ", font=("Inter", 15, "bold"), text_color="#93a4c7").pack(side="left")
        ctk.CTkLabel(
            formula,
            text=f"{money(treasury_total)}",
            font=("Consolas", 16, "bold"),
            text_color=ModernTheme.SUCCESS,
        ).pack(side="left")
        ctk.CTkLabel(formula, text="  +  ", font=("Inter", 15, "bold"), text_color="#93a4c7").pack(side="left")
        ctk.CTkLabel(
            formula,
            text=f"{money(delinquency_total)}",
            font=("Consolas", 16, "bold"),
            text_color=ModernTheme.DANGER,
        ).pack(side="left")
        ctk.CTkLabel(
            formula,
            text=f"Variance: {money(variance)}",
            font=("Consolas", 13, "bold"),
            text_color=status_color,
        ).pack(side="right")

        details = ctk.CTkFrame(self.recon_content, fg_color="transparent")
        details.pack(fill="x", pady=(0, 14))
        rows = (
            ("Equation substitution (A = B + C)", f"{money(total_collectible)} = {money(treasury_total)} + {money(delinquency_total)}"),
            ("Equation variance [A - (B + C)]", money(equation_variance)),
            ("Net current collectible", f"{money(current_levy)} + {money(current_penalty)} - {money(current_discount)} = {money(current_net)}"),
            ("Collections / credits applied", f"{money(calendar_applicable_collections)} + {money(prepaid_current_year)} = {money(treasury_total)}"),
            ("Future-year prepayments excluded from equation", money(future_year_prepayments)),
            ("Prior-year + current-year receivables", f"{money(prior_receivable)} + {money(current_receivable)} = {money(delinquency_total)}"),
            ("As-of tracker variance [tracker - expected ending]", money(tracker_variance)),
            ("All-time vs as-of difference", money(raw_tracker_variance)),
            ("Collection rate", f"{collection_rate:.1f}%"),
            ("Ending receivable rate", f"{delinquency_rate:.1f}%"),
        )
        for label, value in rows:
            row = ctk.CTkFrame(details, fg_color="transparent")
            row.pack(fill="x", pady=2)
            ctk.CTkLabel(row, text=label, font=("Inter", 11), text_color="#bfdbfe", anchor="w").pack(side="left")
            ctk.CTkLabel(row, text=value, font=("Consolas", 11, "bold"), text_color="#f8fafc", anchor="e").pack(side="right")
            ctk.CTkFrame(details, height=1, fg_color="#1f2937").pack(fill="x")

        result_color = "#10b981" if balanced else "#f59e0b"
        result_bg = "#0f302c" if balanced else "#30240f"
        result = ctk.CTkFrame(self.recon_content, fg_color=result_bg, corner_radius=8, border_width=1, border_color=result_color)
        result.pack(fill="x", pady=(0, 14))
        result_text = "Balanced - ready for Accounting review" if balanced else "Needs review before submission"
        detail_text = (
            "The accounting equation and RPT tracker agree within the system tolerance."
            if balanced else
            "Review the diagnostic rows before certification. Timing rows are usually advance or late-posted payments; link gaps and credit rows need data cleanup before Accounting/COA reporting."
        )
        ctk.CTkLabel(result, text=result_text, font=("Inter", 14, "bold"), text_color=result_color).pack(anchor="w", padx=18, pady=(16, 4))
        ctk.CTkLabel(result, text=f"Variance: {money(variance)}", font=("Consolas", 18, "bold"), text_color=result_color).pack(anchor="w", padx=18)
        ctk.CTkLabel(result, text=detail_text, font=("Inter", 11), text_color="#cbd5e1", wraplength=900, justify="left").pack(anchor="w", padx=18, pady=(6, 16))

        diag = ctk.CTkFrame(self.recon_content, fg_color=("#e2e8f0", "#111827"), corner_radius=8, border_width=1, border_color=("#cbd5e1", "#243244"))
        diag.pack(fill="x", pady=(0, 14))
        ctk.CTkLabel(diag, text="RECONCILIATION DIAGNOSTIC", font=("Inter", 11, "bold"), text_color=ModernTheme.TEXT_GRAY).pack(anchor="w", padx=16, pady=(14, 4))
        ctk.CTkLabel(
            diag,
            text=(
                "Rows and categories that can explain why collectible, collections, and receivable totals do not tie exactly. "
                "All-time vs as-of compares today's raw billing balances against the selected fiscal-year reporting view."
            ),
            font=("Inter", 10),
            text_color="#bfdbfe",
            wraplength=1100,
            justify="left",
        ).pack(anchor="w", padx=16, pady=(0, 10))

        diag_summary = ctk.CTkFrame(diag, fg_color="transparent")
        diag_summary.pack(fill="x", padx=16, pady=(0, 10))
        diag_summary.grid_columnconfigure((0, 1, 2, 3, 4, 5), weight=1)

        def diag_chip(parent, col, title, value, color):
            chip = ctk.CTkFrame(parent, fg_color="#0f172a", corner_radius=7, border_width=1, border_color="#334155")
            chip.grid(row=0, column=col, sticky="nsew", padx=(0 if col == 0 else 4, 0 if col == 5 else 4))
            ctk.CTkLabel(chip, text=title.upper(), font=("Inter", 9, "bold"), text_color=ModernTheme.TEXT_GRAY).pack(anchor="w", padx=10, pady=(8, 2))
            ctk.CTkLabel(chip, text=value, font=("Consolas", 12, "bold"), text_color=color).pack(anchor="w", padx=10, pady=(0, 8))

        diag_chip(diag_summary, 0, "Equation variance", money(equation_variance), status_color)
        diag_chip(diag_summary, 1, "All-time vs as-of", money(raw_tracker_variance), "#f59e0b")
        link_issue_count = len(diagnostics.get("payment_link_mismatches", [])) + len(diagnostics.get("unlinked_payments", []))
        diag_chip(diag_summary, 2, "Payment link issues", str(link_issue_count), "#f59e0b")
        diag_chip(diag_summary, 3, "Overpaid / credits", str(len(diagnostics.get("overpaid_or_credit_rows", []))), "#22c55e")
        diag_chip(diag_summary, 4, "Payment-year gaps", str(diagnostics.get("payment_sequence_gap_count", len(diagnostics.get("payment_sequence_gaps", []))) or 0), "#ef4444")
        diag_chip(diag_summary, 5, "Timing/prepayment groups", str(len(diagnostics.get("prior_year_collections", [])) + len(diagnostics.get("future_year_collections", [])) + len(diagnostics.get("current_year_paid_outside_selected_year", []))), "#38bdf8")

        diag_rows = []
        for item in diagnostics.get("unlinked_payments", [])[:10]:
            payment_date = item.get("payment_date") or "No date"
            or_number = item.get("or_number") or "No OR"
            scope = f"{item.get('barangay') or '-'} | {payment_date} | OR {or_number}"
            diag_rows.append((item.get("issue"), item.get("td_number"), item.get("tax_year"), scope, item.get("amount", 0)))
        for item in diagnostics.get("payment_link_mismatches", [])[:8]:
            diag_rows.append((item.get("issue"), item.get("td_number"), item.get("tax_year"), item.get("barangay"), item.get("difference", 0)))
        for item in diagnostics.get("overpaid_or_credit_rows", [])[:8]:
            diag_rows.append((item.get("issue"), item.get("td_number"), item.get("tax_year"), item.get("barangay"), item.get("balance", 0)))
        for item in diagnostics.get("payment_sequence_gaps", [])[:20]:
            later_year = item.get("later_paid_year") or "-"
            status = str(item.get("gap_status") or "gap").replace("_", " ").title()
            scope = f"{item.get('barangay') or '-'} | {status} | Later paid: {later_year}"
            diag_rows.append((item.get("issue"), item.get("td_number"), item.get("tax_year"), scope, item.get("outstanding") or 0))
        for item in diagnostics.get("prior_year_collections", []):
            diag_rows.append((item.get("issue"), "Grouped", item.get("tax_year"), f"{item.get('properties', 0):,} properties", item.get("amount", 0)))
        for item in diagnostics.get("future_year_collections", []):
            diag_rows.append((item.get("issue"), "Grouped", item.get("tax_year"), f"{item.get('properties', 0):,} properties", item.get("amount", 0)))
        for item in diagnostics.get("current_year_paid_outside_details", [])[:12]:
            payment_year = item.get("payment_year") or "No date"
            payment_date = item.get("payment_date") or f"Paid in {payment_year}"
            or_number = item.get("or_number") or "No OR"
            scope = f"{item.get('barangay') or '-'} | {payment_date} | OR {or_number}"
            diag_rows.append((item.get("issue"), item.get("td_number"), item.get("tax_year"), scope, item.get("amount", 0)))
        for item in diagnostics.get("current_year_paid_outside_selected_year", [])[:8]:
            payment_year = item.get("payment_year") or "No date"
            diag_rows.append((item.get("issue"), f"Grouped paid in {payment_year}", item.get("tax_year"), f"{item.get('properties', 0):,} properties", item.get("amount", 0)))
        if not diag_rows:
            for item in diagnostics.get("largest_open_balances", [])[:10]:
                diag_rows.append((item.get("issue"), item.get("td_number"), item.get("tax_year"), item.get("barangay"), item.get("balance", 0)))

        cols = ("Issue", "TD / Group", "Year", "Barangay / Scope", "Amount")
        diag_table = ctk.CTkFrame(diag, fg_color="transparent")
        diag_table.pack(fill="x", padx=16, pady=(0, 14))

        diag_scroll_y = ttk.Scrollbar(diag_table, orient="vertical")
        diag_scroll_x = ttk.Scrollbar(diag_table, orient="horizontal")
        diag_tree = ttk.Treeview(
            diag_table,
            columns=cols,
            show="headings",
            height=min(7, max(3, len(diag_rows))),
            yscrollcommand=diag_scroll_y.set,
            xscrollcommand=diag_scroll_x.set,
        )
        diag_scroll_y.configure(command=diag_tree.yview)
        diag_scroll_x.configure(command=diag_tree.xview)

        widths = {"Issue": 430, "TD / Group": 170, "Year": 90, "Barangay / Scope": 360, "Amount": 150}
        for col in cols:
            diag_tree.heading(col, text=col.upper())
            diag_tree.column(col, width=widths[col], minwidth=widths[col], anchor="e" if col == "Amount" else "center", stretch=False)
        diag_tree.column("Issue", anchor="w")
        diag_tree.column("Barangay / Scope", anchor="w")
        for row in diag_rows:
            diag_tree.insert("", "end", values=(row[0], row[1], row[2], row[3], money(row[4])))

        diag_scroll_y.pack(side="right", fill="y")
        diag_tree.pack(side="top", fill="x", expand=True)
        diag_scroll_x.pack(side="bottom", fill="x")
        if brgy_rows:
            review = ctk.CTkFrame(self.recon_content, fg_color=("#e2e8f0", "#111827"), corner_radius=8, border_width=1, border_color=("#cbd5e1", "#243244"))
            review.pack(fill="both", expand=True)
            ctk.CTkLabel(review, text="TOP BARANGAYS TO REVIEW", font=("Inter", 11, "bold"), text_color=ModernTheme.TEXT_GRAY).pack(anchor="w", padx=16, pady=(14, 8))
            cols = ("Barangay", "Tax Due", "Collected", "Receivable", "Rate")
            tree = ttk.Treeview(review, columns=cols, show="headings", height=6)
            for col in cols:
                tree.heading(col, text=col.upper())
                tree.column(col, anchor="center", width=120)
            tree.column("Barangay", anchor="w", width=200)
            tree.pack(fill="both", expand=True, padx=16, pady=(0, 14))
            top_rows = sorted(brgy_rows, key=lambda row: float(row[6] or 0), reverse=True)[:6]
            for idx, row in enumerate(top_rows):
                due = float(row[2] or 0)
                collected = float(row[5] or 0)
                receivable = float(row[6] or 0)
                rate = (collected / due * 100) if due > 0 else 0
                tree.insert("", "end", values=(row[0], money(due), money(collected), money(receivable), f"{rate:.1f}%"))

    def _recon_money(self, value):
        return f"P {float(value or 0):,.2f}"

    def _build_reconciliation_diag_rows(self, diagnostics, full=False):
        diagnostics = diagnostics or {}
        diag_rows = []

        def take(items, limit):
            items = items or []
            return items if full else items[:limit]

        for item in take(diagnostics.get("unlinked_payments"), 10):
            payment_date = item.get("payment_date") or "No date"
            or_number = item.get("or_number") or "No OR"
            scope = f"{item.get('barangay') or '-'} | {payment_date} | OR {or_number}"
            diag_rows.append((item.get("issue"), item.get("td_number"), item.get("tax_year"), scope, item.get("amount", 0)))
        for item in take(diagnostics.get("payment_link_mismatches"), 8):
            diag_rows.append((item.get("issue"), item.get("td_number"), item.get("tax_year"), item.get("barangay"), item.get("difference", 0)))
        for item in take(diagnostics.get("overpaid_or_credit_rows"), 8):
            diag_rows.append((item.get("issue"), item.get("td_number"), item.get("tax_year"), item.get("barangay"), item.get("balance", 0)))
        for item in take(diagnostics.get("payment_sequence_gaps"), 20):
            later_year = item.get("later_paid_year") or "-"
            status = str(item.get("gap_status") or "gap").replace("_", " ").title()
            scope = f"{item.get('barangay') or '-'} | {status} | Later paid: {later_year}"
            diag_rows.append((item.get("issue"), item.get("td_number"), item.get("tax_year"), scope, item.get("outstanding") or 0))
        for item in diagnostics.get("prior_year_collections", []) or []:
            diag_rows.append((item.get("issue"), "Grouped", item.get("tax_year"), f"{item.get('properties', 0):,} properties", item.get("amount", 0)))
        for item in diagnostics.get("future_year_collections", []) or []:
            diag_rows.append((item.get("issue"), "Grouped", item.get("tax_year"), f"{item.get('properties', 0):,} properties", item.get("amount", 0)))
        for item in take(diagnostics.get("current_year_paid_outside_details"), 12):
            payment_year = item.get("payment_year") or "No date"
            payment_date = item.get("payment_date") or f"Paid in {payment_year}"
            or_number = item.get("or_number") or "No OR"
            scope = f"{item.get('barangay') or '-'} | {payment_date} | OR {or_number}"
            diag_rows.append((item.get("issue"), item.get("td_number"), item.get("tax_year"), scope, item.get("amount", 0)))
        for item in take(diagnostics.get("current_year_paid_outside_selected_year"), 8):
            payment_year = item.get("payment_year") or "No date"
            diag_rows.append((item.get("issue"), f"Grouped paid in {payment_year}", item.get("tax_year"), f"{item.get('properties', 0):,} properties", item.get("amount", 0)))
        if not diag_rows:
            for item in take(diagnostics.get("largest_open_balances"), 10):
                diag_rows.append((item.get("issue"), item.get("td_number"), item.get("tax_year"), item.get("barangay"), item.get("balance", 0)))
        return diag_rows

    def _build_reconciliation_payload(self, data, brgy_rows, metrics, diagnostics, full_diagnostics=False):
        data = data or {}
        brgy_rows = brgy_rows or []
        metrics = metrics or {}
        diagnostics = diagnostics or {}
        assessor = metrics.get("assessor", {})
        treasury = metrics.get("treasury", {})
        delinquency = metrics.get("delinquency", {})

        year = data.get("report_year", self.recon_year_cb.get())
        beginning = float(data.get("beginning_receivable", 0) or 0)
        adjustments = float(data.get("adjustments", 0) or 0)
        collections = float(data.get("collections", 0) or 0)
        current_levy = float(assessor.get("current_year_levy", data.get("current_year_levy", data.get("current_year_assessment", 0))) or 0)
        current_penalty = float(assessor.get("current_year_penalty", data.get("current_year_penalty", 0)) or 0)
        current_discount = float(assessor.get("current_year_discount", data.get("current_year_discount", 0)) or 0)
        current_net = float(assessor.get("current_year_net_collectible", data.get("current_year_net_collectible", data.get("current_year_assessment", 0))) or 0)
        total_collectible = beginning + current_net + adjustments

        calendar_applicable_collections = float(treasury.get("calendar_applicable_collections", data.get("calendar_applicable_collections", collections)) or 0)
        cash_collected_this_year = float(treasury.get("cash_collected_this_year", collections) or 0)
        prepaid_current_year = float(treasury.get("prepaid_current_year", data.get("prepaid_current_year", 0)) or 0)
        future_year_prepayments = float(treasury.get("future_year_prepayments", data.get("future_year_prepayments", 0)) or 0)
        treasury_total = float(treasury.get("total_collected", data.get("applied_collections", calendar_applicable_collections + prepaid_current_year)) or 0)

        expected_unpaid = float(data.get("expected_ending_receivable", data.get("ending_receivable", 0)) or 0)
        tracker_total = float(delinquency.get("total_unpaid", data.get("ending_receivable", 0)) or 0)
        current_receivable = float(delinquency.get("current_year_receivables", 0) or 0)
        prior_receivable = float(delinquency.get("prior_year_receivables", max(tracker_total - current_receivable, 0)) or 0)
        delinquency_total = float(delinquency.get("total_unpaid", prior_receivable + current_receivable) or 0)

        equation_variance = total_collectible - (treasury_total + delinquency_total)
        tracker_variance = float(diagnostics.get("tracker_variance", tracker_total - expected_unpaid) or 0)
        raw_tracker_variance = float(diagnostics.get("raw_tracker_variance", tracker_variance) or 0)
        variance = equation_variance if abs(equation_variance) > 1.0 else tracker_variance
        balanced = max(abs(equation_variance), abs(tracker_variance)) <= 1.0

        top_barangays = []
        for row in sorted(brgy_rows, key=lambda item: float(item[6] or 0), reverse=True)[:20]:
            due = float(row[2] or 0)
            collected = float(row[5] or 0)
            receivable = float(row[6] or 0)
            rate = (collected / due * 100) if due > 0 else 0
            top_barangays.append((row[0], due, collected, receivable, rate))

        return {
            "year": year,
            "prepared_at": datetime.now().strftime("%Y-%m-%d %I:%M %p"),
            "status": "BALANCED" if balanced else "NEEDS REVIEW",
            "balanced": balanced,
            "assessor": {
                "total_assessed_value": float(assessor.get("total_assessed_value", 0) or 0),
                "basic_tax_rate": float(assessor.get("tax_rate_percent", 0) or 0),
                "total_rpt_rate": float(assessor.get("total_tax_rate_percent", 0) or 0),
                "taxable_properties": int(assessor.get("taxable_properties", 0) or 0),
                "current_levy": current_levy,
                "current_penalty": current_penalty,
                "current_discount": current_discount,
                "current_net": current_net,
            },
            "treasury": {
                "basic_tax_collected": float(treasury.get("basic_tax_collected", 0) or 0),
                "cash_collected_this_year": cash_collected_this_year,
                "calendar_applicable_collections": calendar_applicable_collections,
                "prepaid_current_year": prepaid_current_year,
                "future_year_prepayments": future_year_prepayments,
                "accounts_paid": int(treasury.get("accounts_paid", 0) or 0),
                "partial_payments": int(treasury.get("partial_payments", 0) or 0),
                "total_collected": treasury_total,
            },
            "delinquency": {
                "prior_year_receivables": prior_receivable,
                "current_year_receivables": current_receivable,
                "delinquent_accounts": int(delinquency.get("delinquent_accounts", 0) or 0),
                "penalties_interest": float(delinquency.get("penalties_interest", 0) or 0),
                "ending_receivable": delinquency_total,
            },
            "equation": {
                "beginning_receivable": beginning,
                "adjustments": adjustments,
                "total_collectible": total_collectible,
                "collections": treasury_total,
                "ending_receivable": delinquency_total,
                "equation_variance": equation_variance,
                "tracker_variance": tracker_variance,
                "raw_tracker_variance": raw_tracker_variance,
                "variance": variance,
                "collection_rate": (treasury_total / total_collectible * 100) if total_collectible > 0 else 0,
                "receivable_rate": (expected_unpaid / total_collectible * 100) if total_collectible > 0 else 0,
            },
            "diagnostic_counts": {
                "payment_link_issues": len(diagnostics.get("payment_link_mismatches", [])) + len(diagnostics.get("unlinked_payments", [])),
                "overpaid_credits": len(diagnostics.get("overpaid_or_credit_rows", [])),
                "payment_year_gaps": int(diagnostics.get("payment_sequence_gap_count", len(diagnostics.get("payment_sequence_gaps", []))) or 0),
                "timing_prepayment_groups": len(diagnostics.get("prior_year_collections", [])) + len(diagnostics.get("future_year_collections", [])) + len(diagnostics.get("current_year_paid_outside_selected_year", [])),
            },
            "diagnostic_rows": self._build_reconciliation_diag_rows(diagnostics, full=full_diagnostics),
            "top_barangays": top_barangays,
        }

    def _fetch_reconciliation_export_payload(self, year):
        summary = billing.get_rpt_receivables_summary(year)
        metrics = billing.get_reconciliation_metrics(year)
        diagnostics = billing.get_reconciliation_diagnostics(year, limit=500)
        brgy_rows = prop.get_receivables_by_barangay(year=int(year))
        return self._build_reconciliation_payload(
            summary, brgy_rows, metrics, diagnostics, full_diagnostics=True
        )

    def _export_reconciliation_excel(self):
        year = self.recon_year_cb.get()
        dest = filedialog.asksaveasfilename(
            title="Save Reconciliation Working Paper",
            initialfile=f"Reconciliation_Working_Paper_{year}.xlsx",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        )
        if not dest:
            return
        self._run_reconciliation_export(
            self.recon_export_excel_btn,
            "EXPORT EXCEL",
            lambda: self._write_reconciliation_excel(self._fetch_reconciliation_export_payload(year), dest),
            dest,
        )

    def _export_reconciliation_pdf(self):
        year = self.recon_year_cb.get()
        dest = filedialog.asksaveasfilename(
            title="Save Reconciliation PDF",
            initialfile=f"Reconciliation_Working_Paper_{year}.pdf",
            defaultextension=".pdf",
            filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")],
        )
        if not dest:
            return
        self._run_reconciliation_export(
            self.recon_export_pdf_btn,
            "PRINT PDF",
            lambda: self._write_reconciliation_pdf(self._fetch_reconciliation_export_payload(year), dest),
            dest,
        )

    def _run_reconciliation_export(self, button, original_text, worker_fn, dest):
        button.configure(text="GENERATING...", state="disabled")
        self._show_loading()

        def worker():
            try:
                worker_fn()

                def done():
                    self._hide_loading()
                    button.configure(text=original_text, state="normal")
                    if messagebox.askyesno("Export Successful", f"Reconciliation report saved to:\n{dest}\n\nOpen it now?"):
                        self._open_file(dest)

                self.container.after(0, done)
            except Exception as exc:
                self.container.after(
                    0,
                    lambda e=exc: (
                        self._hide_loading(),
                        button.configure(text=original_text, state="normal"),
                        messagebox.showerror("Export Failed", str(e)),
                    ),
                )

        threading.Thread(target=worker, daemon=True).start()

    def _write_reconciliation_excel(self, payload, dest):
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        wb.properties.title = f"Reconciliation Working Paper {payload['year']}"
        wb.properties.subject = "Inter-Department Reconciliation"
        wb.properties.creator = "MTO Treasury System"
        ws = wb.active
        ws.title = "Summary"

        navy = "17365D"
        blue = "1F4E78"
        light_blue = "D9EAF7"
        pale_blue = "EEF5FB"
        dark = "1F2937"
        gray = "E7E6E6"
        light_gray = "F7F9FB"
        green = "008000"
        red = "C00000"
        amber = "B45F06"
        white = "FFFFFF"
        border_color = "A6A6A6"
        thin = Side(style="thin", color=border_color)
        medium = Side(style="medium", color=navy)
        money_fmt = 'P #,##0.00;[Red]-P #,##0.00'
        number_fmt = '#,##0'
        percent_fmt = '0.0%'

        def setup_sheet(sheet, tab_color=blue):
            sheet.sheet_view.showGridLines = False
            sheet.sheet_properties.tabColor = tab_color
            sheet.freeze_panes = "A8"
            sheet.page_setup.orientation = "landscape"
            sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
            sheet.page_margins.left = 0.25
            sheet.page_margins.right = 0.25
            sheet.page_margins.top = 0.5
            sheet.page_margins.bottom = 0.5

        def border_range(sheet, cell_range, side=thin):
            rows = list(sheet[cell_range])
            for row in rows:
                for cell in row:
                    cell.border = Border(top=side, left=side, right=side, bottom=side)

        def fill_range(sheet, cell_range, fill_color):
            fill = PatternFill("solid", fgColor=fill_color)
            for row in sheet[cell_range]:
                for cell in row:
                    cell.fill = fill

        def merge_label(sheet, cell_range, text, fill_color=blue, font_color=white, size=11):
            sheet.merge_cells(cell_range)
            cell = sheet[cell_range.split(":")[0]]
            cell.value = text
            cell.font = Font(name="Calibri", size=size, bold=True, color=font_color)
            cell.fill = PatternFill("solid", fgColor=fill_color)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            border_range(sheet, cell_range, medium)

        def write_value(cell, value, kind="money"):
            cell.value = value
            cell.font = Font(name="Calibri", size=10, bold=True, color=dark)
            cell.alignment = Alignment(horizontal="right", vertical="center")
            if kind == "money":
                cell.number_format = money_fmt
            elif kind == "number":
                cell.number_format = number_fmt
            elif kind == "percent":
                cell.number_format = percent_fmt

        def write_header(sheet, subtitle):
            sheet.merge_cells("A1:H1")
            sheet["A1"] = "REPUBLIC OF THE PHILIPPINES"
            sheet["A1"].font = Font(name="Calibri", size=10, bold=True, color=dark)
            sheet["A1"].alignment = Alignment(horizontal="center")
            sheet.merge_cells("A2:H2")
            sheet["A2"] = "PROVINCE OF AURORA | MUNICIPALITY OF DIPACULAO"
            sheet["A2"].font = Font(name="Calibri", size=10, bold=True, color=dark)
            sheet["A2"].alignment = Alignment(horizontal="center")
            sheet.merge_cells("A3:H3")
            sheet["A3"] = "MUNICIPAL TREASURY OFFICE"
            sheet["A3"].font = Font(name="Calibri", size=11, bold=True, color=dark)
            sheet["A3"].alignment = Alignment(horizontal="center")
            sheet.merge_cells("A4:H4")
            sheet["A4"] = "INTER-DEPARTMENT RECONCILIATION WORKING PAPER"
            sheet["A4"].font = Font(name="Calibri", size=16, bold=True, color=white)
            sheet["A4"].fill = PatternFill("solid", fgColor=navy)
            sheet["A4"].alignment = Alignment(horizontal="center", vertical="center")
            sheet.merge_cells("A5:H5")
            sheet["A5"] = subtitle
            sheet["A5"].font = Font(name="Calibri", size=10, italic=True, color="475569")
            sheet["A5"].alignment = Alignment(horizontal="center")
            sheet.row_dimensions[4].height = 24

        setup_sheet(ws)
        write_header(ws, f"Fiscal Year {payload['year']} | Generated {payload['prepared_at']}")
        for col, width in {
            "A": 28, "B": 18, "C": 3, "D": 30, "E": 18, "F": 3, "G": 30, "H": 18
        }.items():
            ws.column_dimensions[col].width = width

        merge_label(ws, "A7:H7", "CONTROL SUMMARY")
        summary_rows = [
            ("Status", payload["status"], "Variance", payload["equation"]["variance"], "Collection Rate", payload["equation"]["collection_rate"] / 100),
            ("Total Collectible (A)", payload["equation"]["total_collectible"], "Collections / Credits (B)", payload["equation"]["collections"], "Ending Receivable (C)", payload["equation"]["ending_receivable"]),
        ]
        start = 8
        for idx, row in enumerate(summary_rows, start=start):
            ws[f"A{idx}"], ws[f"D{idx}"], ws[f"G{idx}"]
            labels = (("A", row[0]), ("D", row[2]), ("G", row[4]))
            values = (("B", row[1]), ("E", row[3]), ("H", row[5]))
            for col, label in labels:
                ws[f"{col}{idx}"] = label
                ws[f"{col}{idx}"].font = Font(bold=True, color=dark)
                ws[f"{col}{idx}"].fill = PatternFill("solid", fgColor=light_blue)
            for col, value in values:
                kind = "percent" if col == "H" and idx == start else "money"
                if col == "B" and idx == start:
                    ws[f"{col}{idx}"] = value
                    ws[f"{col}{idx}"].font = Font(bold=True, color=green if payload["balanced"] else red)
                    ws[f"{col}{idx}"].alignment = Alignment(horizontal="center", vertical="center")
                else:
                    write_value(ws[f"{col}{idx}"], value, kind)
            border_range(ws, f"A{idx}:H{idx}")

        merge_label(ws, "A12:H12", "RECONCILIATION EQUATION")
        equation_rows = [
            ("Beginning receivable", payload["equation"]["beginning_receivable"]),
            ("Current year net collectible", payload["assessor"]["current_net"]),
            ("Adjustments", payload["equation"]["adjustments"]),
            ("Total collectible (A)", payload["equation"]["total_collectible"]),
            ("Collections / credits (B)", payload["equation"]["collections"]),
            ("Ending receivable (C)", payload["equation"]["ending_receivable"]),
            ("Equation variance [A - (B + C)]", payload["equation"]["equation_variance"]),
            ("As-of tracker variance", payload["equation"]["tracker_variance"]),
            ("All-time vs as-of difference", payload["equation"]["raw_tracker_variance"]),
        ]
        for idx, (label, value) in enumerate(equation_rows, start=13):
            ws.merge_cells(start_row=idx, start_column=1, end_row=idx, end_column=4)
            ws.cell(idx, 1).value = label
            ws.cell(idx, 1).font = Font(color=dark, bold=label.endswith("(A)") or label.endswith("(B)") or label.endswith("(C)"))
            ws.merge_cells(start_row=idx, start_column=5, end_row=idx, end_column=8)
            write_value(ws.cell(idx, 5), value)
            border_range(ws, f"A{idx}:H{idx}")
            if label.startswith("Total collectible") or label.startswith("Equation variance"):
                fill_range(ws, f"A{idx}:H{idx}", pale_blue)

        merge_label(ws, "A24:H24", "DEPARTMENT RECONCILIATION DATA")
        section_headers = [("A25:B25", "ASSESSOR'S OFFICE"), ("D25:E25", "TREASURER'S OFFICE"), ("G25:H25", "RPT TRACKER")]
        for cell_range, title in section_headers:
            merge_label(ws, cell_range, title, fill_color=blue, size=10)
        sections = [
            ("Total assessed value", payload["assessor"]["total_assessed_value"], "money", "Basic tax collected", payload["treasury"]["basic_tax_collected"], "money", "Prior-year receivables", payload["delinquency"]["prior_year_receivables"], "money"),
            ("Basic tax rate", payload["assessor"]["basic_tax_rate"] / 100, "percent", "Cash collected this year", payload["treasury"]["cash_collected_this_year"], "money", "Current year receivables", payload["delinquency"]["current_year_receivables"], "money"),
            ("Total RPT rate", payload["assessor"]["total_rpt_rate"] / 100, "percent", "Current-year prepayments", payload["treasury"]["prepaid_current_year"], "money", "Delinquent accounts", payload["delinquency"]["delinquent_accounts"], "number"),
            ("Taxable properties", payload["assessor"]["taxable_properties"], "number", "Future prepayments excluded", payload["treasury"]["future_year_prepayments"], "money", "Penalties and interest", payload["delinquency"]["penalties_interest"], "money"),
            ("Penalties / interest", payload["assessor"]["current_penalty"], "money", "Accounts paid", payload["treasury"]["accounts_paid"], "number", "Ending receivable", payload["delinquency"]["ending_receivable"], "money"),
            ("Discounts", payload["assessor"]["current_discount"], "money", "Partial payments", payload["treasury"]["partial_payments"], "number", "", "", "text"),
            ("Current year levy", payload["assessor"]["current_levy"], "money", "Applied collections / credits", payload["treasury"]["total_collected"], "money", "", "", "text"),
        ]
        for idx, row in enumerate(sections, start=26):
            for label_col, value_col, label, value, kind in (
                ("A", "B", row[0], row[1], row[2]),
                ("D", "E", row[3], row[4], row[5]),
                ("G", "H", row[6], row[7], row[8]),
            ):
                ws[f"{label_col}{idx}"] = label
                ws[f"{label_col}{idx}"].font = Font(color=dark)
                if kind != "text":
                    write_value(ws[f"{value_col}{idx}"], value, kind)
                else:
                    ws[f"{value_col}{idx}"] = value
            border_range(ws, f"A{idx}:B{idx}")
            border_range(ws, f"D{idx}:E{idx}")
            border_range(ws, f"G{idx}:H{idx}")

        merge_label(ws, "A35:H35", "REVIEW NOTES", fill_color="595959")
        ws.merge_cells("A36:H39")
        ws["A36"] = (
            "This working paper supports reconciliation between assessed collectible amounts, "
            "Treasury collections/credits, and RPT receivable balances. Diagnostic rows should be reviewed "
            "before Accounting or COA submission when the status is NEEDS REVIEW."
        )
        ws["A36"].alignment = Alignment(wrap_text=True, vertical="top")
        ws["A36"].font = Font(color=dark)
        border_range(ws, "A36:H39")

        ws.print_area = "A1:H39"

        diag = wb.create_sheet("Diagnostics")
        setup_sheet(diag, tab_color=amber)
        write_header(diag, "Diagnostic rows that explain timing, credit, or payment-allocation differences")
        for col, width in {"A": 48, "B": 22, "C": 12, "D": 48, "E": 18, "F": 14, "G": 30, "H": 14}.items():
            diag.column_dimensions[col].width = width
        merge_label(diag, "A7:H7", "DIAGNOSTIC SUMMARY", fill_color=amber)
        diag.append([
            "Payment link issues", payload["diagnostic_counts"]["payment_link_issues"],
            "Overpaid / credits", payload["diagnostic_counts"]["overpaid_credits"],
            "Payment-year gaps", payload["diagnostic_counts"]["payment_year_gaps"],
            "Timing / prepayment groups", payload["diagnostic_counts"]["timing_prepayment_groups"],
        ])
        border_range(diag, "A8:H8")
        fill_range(diag, "A8:H8", pale_blue)
        diag.append([])
        diag.append(["Issue", "TD / Group", "Year", "Barangay / Scope", "Amount"])
        for row in payload["diagnostic_rows"]:
            diag.append(list(row))
        for cell in diag[10]:
            cell.font = Font(bold=True, color=white)
            cell.fill = PatternFill("solid", fgColor=navy)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        diag.auto_filter.ref = f"A10:E{max(10, diag.max_row)}"
        diag.freeze_panes = "A11"
        for row in diag.iter_rows(min_row=11):
            row[4].number_format = '#,##0.00'
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

        brgy = wb.create_sheet("Barangay Review")
        setup_sheet(brgy, tab_color=green)
        write_header(brgy, "Top barangays sorted by ending receivable")
        for col, width in {"A": 28, "B": 18, "C": 18, "D": 18, "E": 16}.items():
            brgy.column_dimensions[col].width = width
        merge_label(brgy, "A7:E7", "TOP BARANGAYS TO REVIEW", fill_color=green)
        brgy.append(["Barangay", "Tax Due", "Collected", "Receivable", "Collection Rate"])
        for row in payload["top_barangays"]:
            brgy.append(list(row))
        for cell in brgy[8]:
            cell.font = Font(bold=True, color=white)
            cell.fill = PatternFill("solid", fgColor=navy)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        brgy.auto_filter.ref = f"A8:E{max(8, brgy.max_row)}"
        brgy.freeze_panes = "A9"
        for row in brgy.iter_rows(min_row=9):
            for cell in row[1:4]:
                cell.number_format = money_fmt
            row[4].number_format = '0.0%'
            row[4].value = (row[4].value or 0) / 100
            for cell in row:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

        sign = wb.create_sheet("Certification")
        setup_sheet(sign, tab_color="595959")
        write_header(sign, "Certification and sign-off")
        for col, width in {"A": 24, "B": 24, "C": 24, "D": 24, "E": 24, "F": 24, "G": 24, "H": 24}.items():
            sign.column_dimensions[col].width = width
        merge_label(sign, "A7:H7", "CERTIFICATION", fill_color="595959")
        sign.merge_cells("A9:H11")
        sign["A9"] = (
            "I certify that this reconciliation working paper was generated from the MTO Treasury System "
            "for review by the Municipal Treasury Office, Accounting Office, and COA audit requirements."
        )
        sign["A9"].alignment = Alignment(wrap_text=True, vertical="top")
        sign["A9"].font = Font(color=dark)
        border_range(sign, "A9:H11")
        signature_rows = [
            (14, "Prepared by", "Reviewed by", "Noted by"),
            (18, "Date", "Date", "Date"),
        ]
        for row_idx, left, mid, right in signature_rows:
            for start_col, label in ((1, left), (4, mid), (7, right)):
                end_col = start_col + 1
                sign.merge_cells(start_row=row_idx, start_column=start_col, end_row=row_idx, end_column=end_col)
                sign.cell(row_idx, start_col).value = label
                sign.cell(row_idx, start_col).font = Font(bold=True, color=dark)
                sign.cell(row_idx, start_col).alignment = Alignment(horizontal="center")
                sign.merge_cells(start_row=row_idx + 1, start_column=start_col, end_row=row_idx + 1, end_column=end_col)
                border_range(sign, f"{get_column_letter(start_col)}{row_idx + 1}:{get_column_letter(end_col)}{row_idx + 1}", medium)
        merge_label(sign, "A22:H22", "REMARKS", fill_color="595959")
        sign.merge_cells("A23:H29")
        border_range(sign, "A23:H29")

        for sheet in wb.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(
                        horizontal=cell.alignment.horizontal or "left",
                        vertical=cell.alignment.vertical or "center",
                        wrap_text=cell.alignment.wrap_text,
                    )
                    if cell.font == Font():
                        cell.font = Font(name="Calibri", size=10, color=dark)

        wb.save(dest)
        return dest

    def _write_reconciliation_pdf(self, payload, dest):
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import mm
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

        doc = SimpleDocTemplate(dest, pagesize=landscape(A4), rightMargin=12 * mm, leftMargin=12 * mm, topMargin=10 * mm, bottomMargin=10 * mm)
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle("ReconTitle", parent=styles["Title"], fontName="Helvetica-Bold", fontSize=15, textColor=colors.HexColor("#1e40af"), spaceAfter=4)
        small = ParagraphStyle("Small", parent=styles["BodyText"], fontSize=8, leading=10)
        header = ParagraphStyle("Header", parent=styles["Heading2"], fontSize=10, textColor=colors.HexColor("#0f172a"), spaceBefore=6, spaceAfter=4)

        def money(value):
            return self._recon_money(value)

        story = [
            Paragraph("INTER-DEPARTMENT RECONCILIATION WORKING PAPER", title_style),
            Paragraph(f"Fiscal Year {payload['year']} | Generated {payload['prepared_at']} | Status: <b>{payload['status']}</b>", small),
            Spacer(1, 5),
        ]

        summary_rows = [
            ["Total Collectible (A)", money(payload["equation"]["total_collectible"]), "Collections / Credits (B)", money(payload["equation"]["collections"]), "Ending Receivable (C)", money(payload["equation"]["ending_receivable"])],
            ["Equation", f"{money(payload['equation']['total_collectible'])} = {money(payload['equation']['collections'])} + {money(payload['equation']['ending_receivable'])}", "Variance", money(payload["equation"]["variance"]), "Review Status", payload["status"]],
        ]
        story.append(Table(summary_rows, colWidths=[38 * mm, 42 * mm, 38 * mm, 42 * mm, 38 * mm, 42 * mm], style=TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e2e8f0")),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#94a3b8")),
            ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ])))
        story.append(Spacer(1, 6))

        story.append(Paragraph("Department Data", header))
        department_rows = [
            ["Assessor's Office", "Value", "Treasurer's Office", "Value", "RPT Tracker", "Value"],
            ["Total assessed value", money(payload["assessor"]["total_assessed_value"]), "Basic tax collected", money(payload["treasury"]["basic_tax_collected"]), "Prior-year receivables", money(payload["delinquency"]["prior_year_receivables"])],
            ["Taxable properties", f"{payload['assessor']['taxable_properties']:,}", "Accounts paid", f"{payload['treasury']['accounts_paid']:,}", "Current year receivables", money(payload["delinquency"]["current_year_receivables"])],
            ["Current year levy", money(payload["assessor"]["current_levy"]), "Applied collections / credits", money(payload["treasury"]["total_collected"]), "Ending receivable", money(payload["delinquency"]["ending_receivable"])],
            ["Penalties / discounts", f"{money(payload['assessor']['current_penalty'])} / -{money(payload['assessor']['current_discount'])}", "Future prepayments excluded", money(payload["treasury"]["future_year_prepayments"]), "Delinquent accounts", f"{payload['delinquency']['delinquent_accounts']:,}"],
        ]
        story.append(Table(department_rows, colWidths=[35 * mm, 42 * mm, 38 * mm, 42 * mm, 38 * mm, 42 * mm], repeatRows=1, style=TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e40af")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#94a3b8")),
            ("FONTSIZE", (0, 0), (-1, -1), 7.5),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ])))
        story.append(Spacer(1, 6))

        story.append(Paragraph("Diagnostic Summary", header))
        diag_count_rows = [
            ["Payment link issues", payload["diagnostic_counts"]["payment_link_issues"], "Overpaid / credits", payload["diagnostic_counts"]["overpaid_credits"]],
            ["Payment-year gaps", payload["diagnostic_counts"]["payment_year_gaps"], "Timing / prepayment groups", payload["diagnostic_counts"]["timing_prepayment_groups"]],
        ]
        story.append(Table(diag_count_rows, colWidths=[58 * mm, 24 * mm, 58 * mm, 24 * mm], style=TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#94a3b8")),
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f8fafc")),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
        ])))
        story.append(Spacer(1, 4))

        diag_rows = [["Issue", "TD / Group", "Year", "Barangay / Scope", "Amount"]]
        for issue, td, tax_year, scope, amount in payload["diagnostic_rows"][:40]:
            diag_rows.append([
                Paragraph(str(issue or ""), small),
                str(td or ""),
                str(tax_year or ""),
                Paragraph(str(scope or ""), small),
                money(amount),
            ])
        story.append(Table(diag_rows, colWidths=[70 * mm, 32 * mm, 18 * mm, 70 * mm, 30 * mm], repeatRows=1, style=TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#334155")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#94a3b8")),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ])))
        story.append(Spacer(1, 8))

        story.append(Paragraph("Certification", header))
        sign_rows = [
            ["Prepared by / Date", "", "Reviewed by / Date", "", "Noted by / Date", ""],
            ["", "", "", "", "", ""],
        ]
        story.append(Table(sign_rows, colWidths=[35 * mm, 45 * mm, 35 * mm, 45 * mm, 35 * mm, 45 * mm], rowHeights=[8 * mm, 14 * mm], style=TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#94a3b8")),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ])))
        doc.build(story)
        return dest

    def generate_collection_report(self):
        # Reset to page 1 on a new search
        self._coll_cursors = [None]
        self._coll_page = 0
        self._coll_has_more = False
        self._coll_load_page()

    def _coll_load_page(self):
        month = self.month_cb.get()
        year = self.year_cb.get()
        cursor = self._coll_cursors[self._coll_page]
        self._show_loading()

        def worker():
            try:
                data = billing.get_report_details(
                    month, year,
                    limit=self._coll_page_size,
                    cursor=cursor,
                )
                self.container.after(0, lambda: self._update_coll_table(data))
            except Exception as e:
                self.container.after(
                    0, lambda err=e: messagebox.showerror("Error", str(err))
                )
            finally:
                self.container.after(0, self._hide_loading)

        threading.Thread(target=worker, daemon=True).start()

    def _coll_next_page(self):
        if self._coll_has_more:
            self._coll_page += 1
            self._coll_load_page()

    def _coll_prev_page(self):
        if self._coll_page > 0:
            self._coll_page -= 1
            self._coll_load_page()

    def _update_coll_table(self, data):
        for item in self.coll_tree.get_children():
            self.coll_tree.delete(item)

        # data may be a dict with pagination info or a plain list (legacy)
        if isinstance(data, dict):
            items = data.get("items", [])
            next_cursor = data.get("next_cursor")
            has_more = data.get("has_more", False)
        else:
            items = data or []
            next_cursor = None
            has_more = False

        if not items and self._coll_page == 0:
            self._coll_page_lbl.configure(text="No results")
            self._coll_prev_btn.configure(state="disabled")
            self._coll_next_btn.configure(state="disabled")
            return

        # Store next cursor for the next page
        self._coll_has_more = has_more
        if has_more and next_cursor is not None:
            if len(self._coll_cursors) <= self._coll_page + 1:
                self._coll_cursors.append(next_cursor)
            else:
                self._coll_cursors[self._coll_page + 1] = next_cursor

        for i, row in enumerate(items):
            formatted_row = list(row)
            if len(formatted_row) > 6:
                formatted_row[6] = format_curr(formatted_row[6])
            tag = "evenrow" if i % 2 == 0 else "oddrow"
            self.coll_tree.insert("", "end", values=formatted_row, tags=(tag,))

        # Update pagination controls
        page_num = self._coll_page + 1
        count = len(items)
        self._coll_page_lbl.configure(
            text=f"Page {page_num}  ·  {count} records shown"
        )
        self._coll_prev_btn.configure(
            state="normal" if self._coll_page > 0 else "disabled"
        )
        self._coll_next_btn.configure(
            state="normal" if has_more else "disabled"
        )

    def generate_receivables_report(self):
        year = self.receiv_year_cb.get()
        self._show_loading()

        def worker():
            try:
                data = billing.get_rpt_receivables_summary(year)
                brgy_rows = prop.get_receivables_by_barangay(year=int(year))
                self.container.after(0, lambda: self._update_receiv_summary(data, brgy_rows))
            except Exception as e:
                self.container.after(
                    0, lambda err=e: messagebox.showerror("Error", str(err))
                )
            finally:
                self.container.after(0, self._hide_loading)

        threading.Thread(target=worker, daemon=True).start()

    def _update_receiv_summary(self, data, brgy_rows=None):
        for child in self.receiv_content.winfo_children():
            child.destroy()

        if not data:
            ErrorDialog(self.parent.winfo_toplevel(), tr("reports.tabs.receivables"), tr("reports.errors.no_receivables"))
            return

        # --- DATA PREP ---
        beg = float(data.get("beginning_receivable", 0))
        curr = float(data.get("current_year_net_collectible", data.get("current_year_assessment", 0)))
        curr_levy = float(data.get("current_year_levy", curr))
        curr_penalty = float(data.get("current_year_penalty", 0))
        curr_discount = float(data.get("current_year_discount", 0))
        coll = float(data.get("collections", 0))
        adj = float(data.get("adjustments", 0))
        end = float(data.get("ending_receivable", 0))
        year = data.get("report_year", "N/A")
        brgy_rows = brgy_rows or []

        # Collection efficiency = collections divided by collectible amount.
        total_target = beg + curr + adj
        efficiency = (coll / total_target * 100) if total_target > 0 else 0
        uncollected_rate = (end / total_target * 100) if total_target > 0 else 0

        def money(value):
            return format_curr(value).replace("₱", "P")

        def metric_card(parent, title, value, color, subtitle=None):
            card = ctk.CTkFrame(
                parent,
                fg_color=("#e2e8f0", "#1e293b"),
                corner_radius=8,
                border_width=1,
                border_color=("#cbd5e1", "#334155"),
            )
            ctk.CTkLabel(
                card,
                text=title.upper(),
                font=("Inter", 9, "bold"),
                text_color=ModernTheme.TEXT_GRAY,
            ).pack(anchor="w", padx=14, pady=(12, 2))
            ctk.CTkLabel(
                card,
                text=money(value),
                font=("Inter", 20, "bold"),
                text_color=color,
            ).pack(anchor="w", padx=14)
            if subtitle:
                ctk.CTkLabel(
                    card,
                    text=subtitle,
                    font=("Inter", 10),
                    text_color=ModernTheme.TEXT_GRAY,
                ).pack(anchor="w", padx=14, pady=(2, 12))
            else:
                ctk.CTkFrame(card, height=12, fg_color="transparent").pack()
            return card

        # --- HEADER ---
        header_fr = ctk.CTkFrame(self.receiv_content, fg_color="transparent")
        header_fr.pack(fill="x", pady=(0, 12))
        ctk.CTkLabel(
            header_fr,
            text=f"FISCAL YEAR {year} RECEIVABLES PERFORMANCE",
            font=ModernTheme.H2,
            text_color=ModernTheme.PRIMARY,
        ).pack(side="left")

        eff_color = ModernTheme.SUCCESS if efficiency >= 70 else ModernTheme.WARNING if efficiency >= 40 else ModernTheme.DANGER
        ctk.CTkLabel(
            header_fr,
            text=f"Collection Efficiency: {efficiency:.1f}%",
            font=ModernTheme.H3,
            text_color=eff_color,
        ).pack(side="right")

        formula_fr = ctk.CTkFrame(
            self.receiv_content,
            fg_color=("#dbeafe", "#172033"),
            border_width=1,
            border_color=("#93c5fd", "#334155"),
            corner_radius=8,
        )
        formula_fr.pack(fill="x", pady=(0, 12))
        formula = (
            f"{money(beg)} beginning + {money(curr_levy)} current levy "
            f"+ {money(curr_penalty)} penalties - {money(curr_discount)} discounts "
            f"+ {money(adj)} adjustments - {money(coll)} collections = {money(end)} ending receivable"
        )
        ctk.CTkLabel(
            formula_fr,
            text=formula,
            font=("Inter", 12, "bold"),
            text_color=("#1e293b", "#cbd5e1"),
        ).pack(anchor="w", padx=14, pady=10)

        grid = ctk.CTkFrame(self.receiv_content, fg_color="transparent")
        grid.pack(fill="x", pady=(0, 12))
        for col in range(3):
            grid.grid_columnconfigure(col, weight=1)

        metric_card(grid, "Beginning Balance", beg, "#94a3b8", "Prior-year unpaid balance").grid(row=0, column=0, sticky="nsew", padx=(0, 8), pady=(0, 8))
        metric_card(grid, "Current Year Levy", curr_levy, ModernTheme.PRIMARY, "Assessed value x RPT rate").grid(row=0, column=1, sticky="nsew", padx=4, pady=(0, 8))
        metric_card(grid, "Net Current Collectible", curr, "#f59e0b", "Levy + penalties - discounts").grid(row=0, column=2, sticky="nsew", padx=(8, 0), pady=(0, 8))
        metric_card(grid, "Collections", coll, ModernTheme.SUCCESS, f"{efficiency:.1f}% of collectible amount").grid(row=1, column=0, sticky="nsew", padx=(0, 8), pady=(0, 8))
        metric_card(grid, "Penalties / Interest", curr_penalty, "#f97316", "Shown separately from levy").grid(row=1, column=1, sticky="nsew", padx=4, pady=(0, 8))
        metric_card(grid, "Discounts", curr_discount, "#22c55e", "Deducted from collectible").grid(row=1, column=2, sticky="nsew", padx=(8, 0), pady=(0, 8))
        metric_card(grid, "Adjustments", adj, "#a855f7", "Corrections and adjustments").grid(row=2, column=0, sticky="nsew", padx=(0, 8))
        metric_card(grid, "Total Collectible", total_target, "#f59e0b", "Beginning + net current + adjustments").grid(row=2, column=1, sticky="nsew", padx=4)
        metric_card(grid, "Ending Receivable", end, ModernTheme.DANGER, f"{uncollected_rate:.1f}% remains uncollected").grid(row=2, column=2, sticky="nsew", padx=(8, 0))

        body = ctk.CTkFrame(self.receiv_content, fg_color="transparent")
        body.pack(fill="both", expand=True)
        body.grid_columnconfigure(0, weight=2)
        body.grid_columnconfigure(1, weight=3)
        body.grid_rowconfigure(0, weight=1)

        progress_fr = ctk.CTkFrame(
            body,
            fg_color=("#e2e8f0", "#1e293b"),
            corner_radius=8,
            border_width=1,
            border_color=("#cbd5e1", "#334155"),
        )
        progress_fr.grid(row=0, column=0, sticky="nsew", padx=(0, 8))
        ctk.CTkLabel(
            progress_fr,
            text="COLLECTION TARGET PROGRESS",
            font=("Inter", 11, "bold"),
            text_color=ModernTheme.TEXT_GRAY,
        ).pack(anchor="w", padx=16, pady=(14, 8))

        prog_bar = ctk.CTkProgressBar(progress_fr, height=14, corner_radius=7)
        prog_bar.pack(fill="x", padx=16, pady=(0, 8))
        prog_bar.set(min(1.0, max(0.0, efficiency / 100)))
        prog_bar.configure(progress_color=eff_color)

        remaining = max(0.0, total_target - coll)
        status_text = (
            "Healthy collection pace" if efficiency >= 70
            else "Needs collection follow-up" if efficiency >= 40
            else "Critical collection gap"
        )
        ctk.CTkLabel(
            progress_fr,
            text=f"{status_text}\nRemaining target: {money(remaining)}",
            font=("Inter", 13, "bold"),
            text_color=eff_color,
            justify="left",
        ).pack(anchor="w", padx=16, pady=(6, 4))
        ctk.CTkLabel(
            progress_fr,
            text="Efficiency is collections divided by beginning receivable plus current assessment and adjustments.",
            font=("Inter", 10),
            text_color=ModernTheme.TEXT_GRAY,
            wraplength=430,
            justify="left",
        ).pack(anchor="w", padx=16, pady=(0, 14))

        brgy_fr = ctk.CTkFrame(
            body,
            fg_color=("#e2e8f0", "#1e293b"),
            corner_radius=8,
            border_width=1,
            border_color=("#cbd5e1", "#334155"),
        )
        brgy_fr.grid(row=0, column=1, sticky="nsew", padx=(8, 0))
        ctk.CTkLabel(
            brgy_fr,
            text="TOP RECEIVABLE BARANGAYS",
            font=("Inter", 11, "bold"),
            text_color=ModernTheme.TEXT_GRAY,
        ).pack(anchor="w", padx=16, pady=(14, 8))

        cols = ("Barangay", "Collected", "Receivable", "Rate")
        tree = ttk.Treeview(brgy_fr, columns=cols, show="headings", height=7)
        for col in cols:
            tree.heading(col, text=col.upper())
            tree.column(col, anchor="center", width=110)
        tree.column("Barangay", anchor="w", width=170)
        tree.pack(fill="both", expand=True, padx=16, pady=(0, 14))
        tree.tag_configure("oddrow", background="#1e293b", foreground="#cbd5e1")
        tree.tag_configure("evenrow", background="#162032", foreground="#cbd5e1")

        top_rows = sorted(brgy_rows, key=lambda row: float(row[6] or 0), reverse=True)[:7]
        for idx, row in enumerate(top_rows):
            collected = float(row[5] or 0)
            receivable = float(row[6] or 0)
            due = float(row[2] or 0)
            rate = (collected / due * 100) if due > 0 else 0
            tree.insert(
                "",
                "end",
                values=(row[0], money(collected), money(receivable), f"{rate:.1f}%"),
                tags=("evenrow" if idx % 2 == 0 else "oddrow",),
            )

    def generate_barangay_receivables(self):
        self._show_loading()
        selected = self.brgy_year_cb.get()
        year = None if selected == "All Years" else int(selected)

        def worker():
            try:
                data = prop.get_receivables_by_barangay(year=year)
                self.container.after(0, lambda: self._update_brgy_table(data, year))
            except Exception as e:
                self.container.after(
                    0, lambda err=e: messagebox.showerror("Error", str(err))
                )
            finally:
                self.container.after(0, self._hide_loading)

        threading.Thread(target=worker, daemon=True).start()

    def _update_brgy_table(self, data, year=None):
        for item in self.brgy_tree.get_children():
            self.brgy_tree.delete(item)

        year_label = f"Cumulative as of {year}" if year else "All-time totals"
        self.brgy_year_lbl.configure(text=year_label)

        if not data:
            self.brgy_total_lbl.configure(text=tr("reports.barangay.total").replace("{value}", "P 0.00"))
            return

        grand_total = 0.0
        for i, row in enumerate(data):
            f_row = list(row)
            try:
                receiv_val = float(row[6] or 0)
                grand_total += receiv_val
            except Exception:
                pass

            if len(f_row) >= 7:
                for idx in range(1, 7):
                    f_row[idx] = format_curr(f_row[idx])

            tag = "evenrow" if i % 2 == 0 else "oddrow"
            self.brgy_tree.insert("", "end", values=f_row, tags=(tag,))

        self.brgy_total_lbl.configure(
            text=tr("reports.barangay.total").replace("{value}", f"P {grand_total:,.2f}")
        )

    # ── Export helpers ────────────────────────────────────────────────────────

    @staticmethod
    def _open_file(path):
        """Open a file with the default OS application after saving."""
        try:
            import subprocess, sys
            if sys.platform.startswith("win"):
                os.startfile(path)
            elif sys.platform == "darwin":
                subprocess.call(["open", path])
            else:
                subprocess.call(["xdg-open", path])
        except Exception:
            pass

    def _export_with_feedback(self, btn, worker_fn):
        """Run worker_fn in a thread; show progress on btn, then open the result."""
        original_text = btn.cget("text")
        btn.configure(text="⏳ Generating...", state="disabled")

        def _run():
            try:
                path = worker_fn()
                # Ask where to save
                def _save():
                    dest = filedialog.asksaveasfilename(
                        title="Save Report",
                        initialfile=os.path.basename(path),
                        defaultextension=os.path.splitext(path)[1],
                        filetypes=[
                            ("PDF files", "*.pdf"),
                            ("Excel files", "*.xlsx"),
                            ("All files", "*.*"),
                        ],
                    )
                    if dest:
                        shutil.copy2(path, dest)
                        if messagebox.askyesno("Export Successful",
                                               f"Report saved to:\n{dest}\n\nOpen it now?"):
                            self._open_file(dest)
                    btn.configure(text=original_text, state="normal")

                self.container.after(0, _save)
            except Exception as exc:
                self.container.after(
                    0, lambda e=exc: (
                        messagebox.showerror("Export Failed", str(e)),
                        btn.configure(text=original_text, state="normal"),
                    )
                )

        threading.Thread(target=_run, daemon=True).start()

    def _export_receivables_excel(self):
        year = self.receiv_year_cb.get()
        self._export_with_feedback(
            self.receiv_export_excel_btn,
            lambda: billing.export_report_excel("receivables", year=year),
        )

    def _export_receivables_barangay_pdf(self):
        year = int(self.receiv_year_cb.get())
        self._export_with_feedback(
            self.receiv_export_brgy_btn,
            lambda: billing.download_receivables_by_barangay_pdf(year=year),
        )

    def _export_brgy_pdf(self):
        selected = self.brgy_year_cb.get()
        year = None if selected == "All Years" else int(selected)
        btn = None
        # Walk the widget tree to find our PDF button
        try:
            tab_children = self.barangay_tab.winfo_children()
            for w in tab_children:
                for w2 in w.winfo_children():
                    for w3 in w2.winfo_children():
                        if hasattr(w3, "cget") and "Export PDF" in str(w3.cget("text")):
                            btn = w3
        except Exception:
            pass
        if btn is None:
            btn = ctk.CTkButton(self.barangay_tab, text="")

        self._export_with_feedback(
            btn,
            lambda: billing.download_receivables_by_barangay_pdf(year=year)
        )

    def _export_brgy_excel(self):
        selected = self.brgy_year_cb.get()
        year_str = selected if selected != "All Years" else "All"
        btn = None
        try:
            tab_children = self.barangay_tab.winfo_children()
            for w in tab_children:
                for w2 in w.winfo_children():
                    for w3 in w2.winfo_children():
                        if hasattr(w3, "cget") and "Export Excel" in str(w3.cget("text")):
                            btn = w3
        except Exception:
            pass
        if btn is None:
            btn = ctk.CTkButton(self.barangay_tab, text="")

        self._export_with_feedback(
            btn,
            lambda: billing.export_report_excel("receivables_by_barangay", year=year_str),
        )

    def open_manage_deposits(self):
        start_date, end_date = self.get_selected_date_range()
        ManageDepositsModal(self.container, start_date, end_date)

    def open_export_signatories(self):
        default_officer = self.user.get("full_name", "") if isinstance(self.user, dict) else ""
        SignatoriesModal(self.container, default_officer, self.export_coa_rcd)

    def export_coa_rcd(self, liquidating_officer, treasurer):
        start_date, end_date = self.get_selected_date_range()
        self._show_loading()
        
        self.export_rcd_btn.configure(state="disabled", text="⏳ Exporting...")
        
        def worker():
            try:
                path = reports_api.download_coa_rcd(
                    start_date=start_date,
                    end_date=end_date,
                    liquidating_officer=liquidating_officer,
                    treasurer=treasurer
                )
                
                def save_file():
                    filename = f"COA_RCD_{start_date}_to_{end_date}.xlsx"
                    dest = filedialog.asksaveasfilename(
                        title="Save COA RCD Report",
                        initialfile=filename,
                        defaultextension=".xlsx",
                        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
                    )
                    if dest:
                        shutil.copy2(path, dest)
                        if messagebox.askyesno("Export Successful", f"COA RCD report saved to:\n{dest}\n\nOpen it now?"):
                            self._open_file(dest)
                            
                self.container.after(0, save_file)
            except Exception as e:
                self.container.after(0, lambda: messagebox.showerror("Export Failed", str(e)))
            finally:
                self.container.after(0, lambda: [
                    self._hide_loading(),
                    self.export_rcd_btn.configure(state="normal", text="📊 Export COA RCD (Excel)")
                ])
                
        threading.Thread(target=worker, daemon=True).start()

    def get_selected_date_range(self):
        month = self.month_cb.get()
        year = self.year_cb.get()
        
        import calendar
        now = datetime.now()
        y = int(year) if year != "All" else now.year
        
        if month == "All":
            start_date = f"{y}-01-01"
            end_date = f"{y}-12-31"
        else:
            m = int(month)
            _, last_day = calendar.monthrange(y, m)
            start_date = f"{y}-{m:02d}-01"
            end_date = f"{y}-{m:02d}-{last_day:02d}"
            
        return start_date, end_date


class SignatoriesModal(ctk.CTkToplevel):
    def __init__(self, parent, default_officer, callback):
        super().__init__(parent)
        self.title("RCD Signatories")
        self.geometry("400x320")
        self.resizable(False, False)
        self.callback = callback
        
        self.transient(parent.winfo_toplevel())
        self.grab_set()
        self.attributes("-topmost", True)
        
        self.update_idletasks()
        sw, sh = self.winfo_screenwidth(), self.winfo_screenheight()
        self.geometry(f"+{(sw-400)//2}+{(sh-320)//2}")
        
        self.setup_ui(default_officer)
        
    def setup_ui(self, default_officer):
        self.configure(fg_color="#1a1a2e")
        ctk.CTkFrame(self, fg_color=ModernTheme.PRIMARY, height=4, corner_radius=0).pack(fill="x")
        
        body = ctk.CTkFrame(self, fg_color="#1a1a2e")
        body.pack(fill="both", expand=True, padx=30, pady=24)
        
        ctk.CTkLabel(body, text="COA RCD Signatories", font=("Segoe UI", 14, "bold"), text_color="white").pack(anchor="w", pady=(0, 6))
        ctk.CTkLabel(body, text="Enter the names of the signing officers for this report.", font=("Segoe UI", 11), text_color="#a0aec0").pack(anchor="w", pady=(0, 16))
        
        ctk.CTkLabel(body, text="Liquidating Officer", font=("Segoe UI", 11, "bold"), text_color="#a0aec0").pack(anchor="w")
        self.officer_ent = ctk.CTkEntry(body, height=36, fg_color="#2d2d4e", border_color="#4a4a6e", text_color="white")
        self.officer_ent.insert(0, default_officer)
        self.officer_ent.pack(fill="x", pady=(4, 12))
        self.officer_ent.bind("<Return>", lambda _e: self.on_export())
        self.officer_ent.bind("<KP_Enter>", lambda _e: self.on_export())
        
        ctk.CTkLabel(body, text="Municipal Treasurer", font=("Segoe UI", 11, "bold"), text_color="#a0aec0").pack(anchor="w")
        self.treasurer_ent = ctk.CTkEntry(body, height=36, fg_color="#2d2d4e", border_color="#4a4a6e", text_color="white")
        self.treasurer_ent.pack(fill="x", pady=(4, 16))
        self.treasurer_ent.bind("<Return>", lambda _e: self.on_export())
        self.treasurer_ent.bind("<KP_Enter>", lambda _e: self.on_export())
        
        btn_fr = ctk.CTkFrame(body, fg_color="transparent")
        btn_fr.pack(fill="x")
        
        ctk.CTkButton(btn_fr, text="Cancel", command=self.destroy,
                      fg_color="#2d2d4e", hover_color="#3d3d5e", text_color="#a0aec0",
                      height=38, width=100, font=("Segoe UI", 11, "bold"), corner_radius=8).pack(side="left")
                      
        ctk.CTkButton(btn_fr, text="Export Excel", command=self.on_export,
                      fg_color=ModernTheme.SUCCESS, hover_color="#1e7e34", text_color="white",
                      height=38, width=150, font=("Segoe UI", 11, "bold"), corner_radius=8).pack(side="right")
                      
    def on_export(self):
        officer = self.officer_ent.get().strip()
        treasurer = self.treasurer_ent.get().strip()
        if not officer or not treasurer:
            messagebox.showerror("Error", "Both fields are required.")
            return
        self.callback(officer, treasurer)
        self.destroy()


class ManageDepositsModal(ctk.CTkToplevel):
    def __init__(self, parent, start_date: str, end_date: str):
        super().__init__(parent)
        self.title("Manage Bank Deposits")
        self.geometry("750x550")
        self.resizable(False, False)
        self.start_date = start_date
        self.end_date = end_date
        
        self.transient(parent.winfo_toplevel())
        self.grab_set()
        self.attributes("-topmost", True)
        
        self.update_idletasks()
        sw, sh = self.winfo_screenwidth(), self.winfo_screenheight()
        self.geometry(f"+{(sw-750)//2}+{(sh-550)//2}")
        
        self.setup_ui()
        self.refresh_deposits()
        
    def setup_ui(self):
        self.configure(fg_color="#1a1a2e")
        ctk.CTkFrame(self, fg_color=ModernTheme.PRIMARY, height=4, corner_radius=0).pack(fill="x")
        
        main_fr = ctk.CTkFrame(self, fg_color="transparent")
        main_fr.pack(fill="both", expand=True, padx=20, pady=20)
        
        title_fr = ctk.CTkFrame(main_fr, fg_color="transparent")
        title_fr.pack(fill="x", pady=(0, 10))
        ctk.CTkLabel(title_fr, text="🏦 Bank Deposit Slips", font=("Segoe UI", 16, "bold"), text_color="white").pack(anchor="w")
        ctk.CTkLabel(title_fr, text=f"Date Range: {self.start_date} to {self.end_date}", font=("Segoe UI", 11), text_color="#a0aec0").pack(anchor="w")
        
        content_fr = ctk.CTkFrame(main_fr, fg_color="transparent")
        content_fr.pack(fill="both", expand=True, pady=10)
        content_fr.grid_columnconfigure(0, weight=1)
        content_fr.grid_columnconfigure(1, weight=2)
        content_fr.grid_rowconfigure(0, weight=1)
        
        # --- LEFT: FORM ---
        form_fr = ctk.CTkFrame(content_fr, fg_color="#252538", corner_radius=8)
        form_fr.grid(row=0, column=0, sticky="nsew", padx=(0, 10))
        
        ctk.CTkLabel(form_fr, text="Log New Deposit", font=("Segoe UI", 12, "bold"), text_color="white").pack(anchor="w", padx=15, pady=(15, 10))
        
        ctk.CTkLabel(form_fr, text="Date Deposited (YYYY-MM-DD)", font=("Segoe UI", 10), text_color="#a0aec0").pack(anchor="w", padx=15)
        self.date_ent = ctk.CTkEntry(form_fr, height=32, fg_color="#1a1a2e", border_color="#4a4a6e", text_color="white")
        self.date_ent.insert(0, datetime.now().strftime("%Y-%m-%d"))
        self.date_ent.pack(fill="x", padx=15, pady=(4, 10))
        self.date_ent.bind("<Return>", lambda _e: self.save_deposit())
        self.date_ent.bind("<KP_Enter>", lambda _e: self.save_deposit())
        
        ctk.CTkLabel(form_fr, text="Bank Name / Branch", font=("Segoe UI", 10), text_color="#a0aec0").pack(anchor="w", padx=15)
        self.bank_ent = ctk.CTkEntry(form_fr, placeholder_text="e.g. Landbank", height=32, fg_color="#1a1a2e", border_color="#4a4a6e", text_color="white")
        self.bank_ent.pack(fill="x", padx=15, pady=(4, 10))
        self.bank_ent.bind("<Return>", lambda _e: self.save_deposit())
        self.bank_ent.bind("<KP_Enter>", lambda _e: self.save_deposit())
        
        ctk.CTkLabel(form_fr, text="Reference / Slip No.", font=("Segoe UI", 10), text_color="#a0aec0").pack(anchor="w", padx=15)
        self.ref_ent = ctk.CTkEntry(form_fr, placeholder_text="e.g. DS-12345", height=32, fg_color="#1a1a2e", border_color="#4a4a6e", text_color="white")
        self.ref_ent.pack(fill="x", padx=15, pady=(4, 10))
        self.ref_ent.bind("<Return>", lambda _e: self.save_deposit())
        self.ref_ent.bind("<KP_Enter>", lambda _e: self.save_deposit())
        
        ctk.CTkLabel(form_fr, text="Amount (₱)", font=("Segoe UI", 10), text_color="#a0aec0").pack(anchor="w", padx=15)
        self.amt_ent = ctk.CTkEntry(form_fr, placeholder_text="e.g. 50000.00", height=32, fg_color="#1a1a2e", border_color="#4a4a6e", text_color="white")
        self.amt_ent.pack(fill="x", padx=15, pady=(4, 15))
        self.amt_ent.bind("<Return>", lambda _e: self.save_deposit())
        self.amt_ent.bind("<KP_Enter>", lambda _e: self.save_deposit())
        
        ctk.CTkButton(form_fr, text="➕ Save Deposit", command=self.save_deposit,
                      fg_color=ModernTheme.PRIMARY, hover_color="#2c6ea1", text_color="white",
                      height=36, font=("Segoe UI", 11, "bold")).pack(fill="x", padx=15, pady=(0, 15))
        
        # --- RIGHT: TABLE ---
        table_container = ctk.CTkFrame(content_fr, fg_color="#252538", corner_radius=8)
        table_container.grid(row=0, column=1, sticky="nsew", padx=(10, 0))
        
        ctk.CTkLabel(table_container, text="Existing Deposits", font=("Segoe UI", 12, "bold"), text_color="white").pack(anchor="w", padx=15, pady=(15, 10))
        
        tree_fr = ctk.CTkFrame(table_container, fg_color="transparent")
        tree_fr.pack(fill="both", expand=True, padx=15, pady=(0, 10))
        
        scrolly = ttk.Scrollbar(tree_fr, orient="vertical")
        scrolly.pack(side="right", fill="y")
        
        cols = ("ID", "Date", "Bank", "Ref No", "Amount", "Logged By")
        self.tree = ttk.Treeview(tree_fr, columns=cols, show="headings", yscrollcommand=scrolly.set, height=10)
        scrolly.configure(command=self.tree.yview)
        
        self.tree.heading("ID", text="ID")
        self.tree.heading("Date", text="DATE")
        self.tree.heading("Bank", text="BANK")
        self.tree.heading("Ref No", text="REF NO")
        self.tree.heading("Amount", text="AMOUNT")
        self.tree.heading("Logged By", text="BY")
        
        self.tree.column("ID", width=40, anchor="center")
        self.tree.column("Date", width=90, anchor="center")
        self.tree.column("Bank", width=120, anchor="w")
        self.tree.column("Ref No", width=90, anchor="center")
        self.tree.column("Amount", width=100, anchor="right")
        self.tree.column("Logged By", width=80, anchor="center")
        
        self.tree.pack(fill="both", expand=True)
        
        self.tree.tag_configure("oddrow", background="#2b2b2b", foreground="white")
        self.tree.tag_configure("evenrow", background="#333333", foreground="white")
        
        btn_fr = ctk.CTkFrame(table_container, fg_color="transparent")
        btn_fr.pack(fill="x", padx=15, pady=(0, 15))
        
        ctk.CTkButton(btn_fr, text="🗑️ Delete Selected", command=self.delete_deposit,
                      fg_color="#c0392b", hover_color="#962d22", text_color="white",
                      width=150, height=32, font=("Segoe UI", 11, "bold")).pack(side="right")

        close_fr = ctk.CTkFrame(main_fr, fg_color="transparent")
        close_fr.pack(fill="x", pady=(10, 0))
        ctk.CTkButton(close_fr, text="Close", command=self.destroy,
                      fg_color="#2d2d4e", hover_color="#3d3d5e", text_color="#a0aec0",
                      width=100, height=36, font=("Segoe UI", 11, "bold"), corner_radius=8).pack(side="right")
                      
    def refresh_deposits(self):
        for item in self.tree.get_children():
            self.tree.delete(item)
            
        def worker():
            try:
                data = reports_api.list_bank_deposits(self.start_date, self.end_date)
                self.after(0, lambda: self._update_table(data))
            except Exception as e:
                self.after(0, lambda: messagebox.showerror("Error", f"Failed to load deposits: {e}"))
                
        threading.Thread(target=worker, daemon=True).start()
        
    def _update_table(self, data):
        for i, d in enumerate(data):
            tag = "evenrow" if i % 2 == 0 else "oddrow"
            self.tree.insert("", "end", values=(
                d["id"],
                d["date_deposited"],
                d["bank_name"],
                d["reference_number"],
                format_curr(d["amount"]),
                d["deposited_by"]
            ), tags=(tag,))
            
    def save_deposit(self):
        date_val = self.date_ent.get().strip()
        bank_val = self.bank_ent.get().strip()
        ref_val = self.ref_ent.get().strip()
        amt_val = self.amt_ent.get().strip()
        
        if not date_val or not bank_val or not ref_val or not amt_val:
            messagebox.showerror("Error", "All fields are required.")
            return
            
        try:
            amt = float(amt_val)
            if amt <= 0:
                raise ValueError()
        except ValueError:
            messagebox.showerror("Error", "Amount must be a positive number.")
            return
            
        try:
            datetime.strptime(date_val, "%Y-%m-%d")
        except ValueError:
            messagebox.showerror("Error", "Invalid Date format. Use YYYY-MM-DD.")
            return
            
        def worker():
            try:
                reports_api.log_bank_deposit(date_val, bank_val, ref_val, amt)
                self.after(0, lambda: [
                    self.bank_ent.delete(0, tk.END),
                    self.ref_ent.delete(0, tk.END),
                    self.amt_ent.delete(0, tk.END),
                    self.refresh_deposits(),
                    show_toast(self.winfo_toplevel(), "Deposit logged successfully.", type="success")
                ])
            except Exception as e:
                self.after(0, lambda: messagebox.showerror("Error", str(e)))
                
        threading.Thread(target=worker, daemon=True).start()
        
    def delete_deposit(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showerror("Error", "Please select a deposit slip to delete.")
            return
            
        vals = self.tree.item(sel[0])["values"]
        dep_id = int(vals[0])
        bank = vals[2]
        ref = vals[3]
        
        if not messagebox.askyesno("Confirm Delete", f"Are you sure you want to delete deposit Ref {ref} to {bank}?"):
            return
            
        def worker():
            try:
                reports_api.delete_bank_deposit(dep_id)
                self.after(0, lambda: [
                    self.refresh_deposits(),
                    show_toast(self.winfo_toplevel(), "Deposit deleted.", type="info")
                ])
            except Exception as e:
                self.after(0, lambda: messagebox.showerror("Error", str(e)))
                
        threading.Thread(target=worker, daemon=True).start()

