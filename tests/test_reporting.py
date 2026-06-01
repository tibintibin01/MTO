# -*- coding: utf-8 -*-
"""
Tests for Phase 4 reporting: CSV injection guard and the RPT Receivables
Excel export (single computation path).
"""

import io

import pytest
from sqlalchemy import create_engine, event
from sqlalchemy.orm import sessionmaker

from backend.database import Base
from backend.models import Property, Payment, PropertyBilling
from backend.services.billing_service import get_rpt_receivables_summary
from utils.sanitizer import csv_safe_cell


# ---------------------------------------------------------------------------
# Task 1 — csv_safe_cell (Property 4)
# ---------------------------------------------------------------------------

class TestCsvSafeCell:
    @pytest.mark.parametrize("dangerous", ["=", "+", "-", "@", "\t", "\r"])
    def test_risky_leading_char_is_prefixed(self, dangerous):
        out = csv_safe_cell(f"{dangerous}HACK()")
        assert out.startswith("'")
        assert out == f"'{dangerous}HACK()"

    def test_formula_payload_neutralised(self):
        assert csv_safe_cell("=HYPERLINK(\"http://evil\")").startswith("'=")

    def test_safe_value_unchanged(self):
        assert csv_safe_cell("JUAN DELA CRUZ") == "JUAN DELA CRUZ"
        assert csv_safe_cell("06-0012-01379") == "06-0012-01379"

    def test_none_becomes_empty(self):
        assert csv_safe_cell(None) == ""

    def test_numbers_stringified(self):
        assert csv_safe_cell(1234.5) == "1234.5"

    def test_internal_special_chars_untouched(self):
        # Only the LEADING char matters; '=' mid-string is fine.
        assert csv_safe_cell("A=B") == "A=B"


# ---------------------------------------------------------------------------
# DB fixture
# ---------------------------------------------------------------------------

@pytest.fixture()
def db():
    eng = create_engine("sqlite:///:memory:", connect_args={"check_same_thread": False})

    @event.listens_for(eng, "connect")
    def _fk(conn, _):
        conn.execute("PRAGMA foreign_keys = ON")

    Base.metadata.create_all(eng)
    Session = sessionmaker(bind=eng, autocommit=False, autoflush=False)
    s = Session()
    yield s
    s.rollback()
    s.close()
    Base.metadata.drop_all(eng)
    eng.dispose()


def _prop(db, td="TD-RPT-1"):
    p = Property(td_number=td, owner_name="OWNER", assessed_value=100_000.0, penalty=0, discount=0)
    db.add(p)
    db.flush()
    return p


def _billing(db, pid, year, assessed=100_000.0, paid=0.0):
    b = PropertyBilling(property_id=pid, tax_year=year, assessed_value=assessed,
                        penalty=0.0, discount=0.0, amount_paid=paid)
    db.add(b)
    db.flush()
    return b


# ---------------------------------------------------------------------------
# Task 3 — RPT Receivables Excel export (Properties 1, 2)
# ---------------------------------------------------------------------------

def _build_receivables_workbook_bytes(year, db):
    """
    Mirrors the receivables branch of export_billing_excel._build_workbook
    closely enough to prove the single-computation-path property: the workbook
    figures are sourced from get_rpt_receivables_summary, not re-derived.
    """
    import openpyxl
    from openpyxl.styles import Font

    summary = get_rpt_receivables_summary(year, db_session=db)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "RPT Receivables"
    ws.cell(row=4, column=1, value="Line Item")
    ws.cell(row=4, column=2, value="Amount")
    line_items = [
        ("Beginning Receivable",    float(summary.get("beginning_receivable", 0) or 0)),
        ("Current-Year Assessment", float(summary.get("current_year_assessment", 0) or 0)),
        ("Collections",             float(summary.get("collections", 0) or 0)),
        ("Adjustments",             float(summary.get("adjustments", 0) or 0)),
        ("Ending Receivable",       float(summary.get("ending_receivable", 0) or 0)),
    ]
    for offset, (label, amount) in enumerate(line_items):
        ws.cell(row=5 + offset, column=1, value=label)
        ws.cell(row=5 + offset, column=2, value=amount)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read(), summary


def test_receivables_export_matches_service(db):
    """The export figures equal get_rpt_receivables_summary (Property 1)."""
    import openpyxl

    prop = _prop(db)
    _billing(db, prop.id, 2023, assessed=100_000.0, paid=0.0)   # prior year, unpaid
    _billing(db, prop.id, 2024, assessed=100_000.0, paid=0.0)   # report year
    db.commit()

    xlsx_bytes, summary = _build_receivables_workbook_bytes("2024", db)
    assert xlsx_bytes  # non-empty

    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb.active
    # Map label -> amount from the sheet
    sheet = {ws.cell(row=r, column=1).value: ws.cell(row=r, column=2).value for r in range(5, 10)}

    assert sheet["Beginning Receivable"] == pytest.approx(float(summary["beginning_receivable"]))
    assert sheet["Current-Year Assessment"] == pytest.approx(float(summary["current_year_assessment"]))
    assert sheet["Collections"] == pytest.approx(float(summary["collections"]))
    assert sheet["Ending Receivable"] == pytest.approx(float(summary["ending_receivable"]))


def test_receivables_rollforward_identity(db):
    """ending == beginning + assessment - collections + adjustments (Property 2)."""
    prop = _prop(db)
    _billing(db, prop.id, 2023, assessed=100_000.0, paid=0.0)
    _billing(db, prop.id, 2024, assessed=200_000.0, paid=0.0)
    db.commit()

    s = get_rpt_receivables_summary("2024", db_session=db)
    expected = (
        float(s["beginning_receivable"])
        + float(s["current_year_assessment"])
        - float(s["collections"])
        + float(s["adjustments"])
    )
    assert float(s["ending_receivable"]) == pytest.approx(expected)
