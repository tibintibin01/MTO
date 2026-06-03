# -*- coding: utf-8 -*-
import io
import pytest
from datetime import datetime, timezone
from decimal import Decimal
from sqlalchemy import create_engine, event
from sqlalchemy.orm import sessionmaker
import openpyxl

from backend.database import Base
from backend.models import Property, Payment, BankDeposit, User
from backend.services.coa_rcd_generator import is_current_year, generate_coa_rcd_excel

@pytest.fixture()
def db():
    # In-memory SQLite DB for fast unit tests
    eng = create_engine("sqlite:///:memory:", connect_args={"check_same_thread": False})

    @event.listens_for(eng, "connect")
    def _fk(conn, _):
        conn.execute("PRAGMA foreign_keys = ON")

    Base.metadata.create_all(eng)
    Session = sessionmaker(bind=eng, autocommit=False, autoflush=False)
    s = Session()
    yield s
    s.rollback()
    s.close()
    Base.metadata.drop_all(eng)
    eng.dispose()

def test_is_current_year_logic():
    # date_paid in 2026, tax_year is 2026 => Current
    assert is_current_year("2026", datetime(2026, 6, 3)) is True
    # date_paid in 2026, tax_year is 2027 => Current/Advance
    assert is_current_year("2027", datetime(2026, 6, 3)) is True
    # date_paid in 2026, tax_year is 2025 => Prior
    assert is_current_year("2025", datetime(2026, 6, 3)) is False
    # Handle ranged string years
    assert is_current_year("2026-2027", datetime(2026, 6, 3)) is True
    assert is_current_year("2025-2026", datetime(2026, 6, 3)) is False
    # None/invalid fallbacks
    assert is_current_year(None, datetime(2026, 6, 3)) is True
    assert is_current_year("invalid", datetime(2026, 6, 3)) is True

def test_generate_coa_rcd_excel(db):
    # 1. Create property
    prop = Property(
        td_number="06-0012-0001",
        owner_name="JUAN DELA CRUZ",
        payor_name="JUAN DELA CRUZ JR",
        assessed_value=Decimal("150000.00")
    )
    db.add(prop)
    db.flush()

    # 2. Add payments (Date: 2026-06-03)
    # Payment 1: Current year tax (amount=10000, penalty=400, discount=200)
    pay1 = Payment(
        property_id=prop.id,
        amount=Decimal("10000.00"),
        penalty=Decimal("400.00"),
        discount=Decimal("200.00"),
        or_number="OR-0001",
        tax_year="2026",
        date_paid=datetime(2026, 6, 3, 10, 30, 0)
    )
    # Payment 2: Prior year tax (amount=6000, penalty=1200, discount=0)
    pay2 = Payment(
        property_id=prop.id,
        amount=Decimal("6000.00"),
        penalty=Decimal("1200.00"),
        discount=Decimal("0.00"),
        or_number="OR-0002",
        tax_year="2025",
        date_paid=datetime(2026, 6, 3, 14, 15, 0)
    )
    db.add_all([pay1, pay2])
    db.flush()

    # 3. Add bank deposit
    dep = BankDeposit(
        date_deposited=datetime(2026, 6, 3, 16, 0, 0),
        bank_name="Landbank - Dipaculao Branch",
        reference_number="DS-9999",
        amount=Decimal("17400.00"),  # Total cash collected = (10000 + 400 - 200) + (6000 + 1200 - 0) = 10200 + 7200 = 17400
        deposited_by="cashier_kevin"
    )
    db.add(dep)
    db.flush()
    db.commit()

    # 4. Generate report
    start_dt = datetime(2026, 6, 3)
    end_dt = datetime(2026, 6, 3)
    excel_buf = generate_coa_rcd_excel(db, start_dt, end_dt, "KEVIN LIQUIDATOR", "JUAN TREASURER")
    assert excel_buf is not None
    
    # 5. Validate the generated workbook
    wb = openpyxl.load_workbook(io.BytesIO(excel_buf.read()))
    
    # Verify sheet names
    assert "Summary" in wb.sheetnames
    assert "Detailed Collections" in wb.sheetnames
    assert "UACS Summary" in wb.sheetnames
    
    # Check Summary Tab values
    ws_summary = wb["Summary"]
    
    # Check Header info is written
    assert ws_summary["A5"].value == "REPORT OF COLLECTIONS AND DEPOSITS"
    
    # Row 10: Basic Current = 10000 / 2 = 5000
    assert ws_summary["A10"].value == "Real Property Tax - Basic (Current Year)"
    assert ws_summary["B10"].value == "4-01-02-040-01"
    assert ws_summary["C10"].value == 5000.0
    
    # Row 11: Basic Prior = 6000 / 2 = 3000
    assert ws_summary["A11"].value == "Real Property Tax - Basic (Prior Years)"
    assert ws_summary["C11"].value == 3000.0
    
    # Row 12: Basic Penalty = (400 + 1200) / 2 / 2 ?
    # Wait, total penalty is 400 + 1200 = 1600. Split 50/50 is 800 Basic Penalty, 800 SEF Penalty.
    assert ws_summary["A12"].value == "Real Property Tax - Basic (Penalties)"
    assert ws_summary["C12"].value == 800.0
    
    # Row 13: Basic Discount = 200 / 2 = 100
    assert ws_summary["A13"].value == "Discount on Real Property Tax - Basic"
    assert ws_summary["C13"].value == -100.0
    
    # Check Total Collections in C18
    # Total Collections = 17400
    assert ws_summary["C18"].value == 17400.0

    # Check deposits table starts around row 22/23/24 depending on spacing
    # Search for Bank name row
    found_deposit = False
    for r in range(20, 35):
        if ws_summary[f"B{r}"].value == "Landbank - Dipaculao Branch":
            assert ws_summary[f"A{r}"].value == "2026-06-03"
            assert ws_summary[f"C{r}"].value == "DS-9999"
            assert ws_summary[f"D{r}"].value == 17400.0
            found_deposit = True
            break
    assert found_deposit is True

    # Validate Detailed Collections Tab
    ws_detailed = wb["Detailed Collections"]
    assert ws_detailed["C5"].value == "JUAN DELA CRUZ JR"
    assert ws_detailed["D5"].value == 5000.0
    assert ws_detailed["E5"].value == 5000.0
    assert ws_detailed["H5"].value == 10200.0 # 5000 + 5000 + 400 - 200 = 10200
    
    # Validate UACS Summary Tab
    ws_uacs = wb["UACS Summary"]
    uacs_map = {ws_uacs[f"A{r}"].value: ws_uacs[f"C{r}"].value for r in range(5, 13)}
    
    assert uacs_map["4-01-02-040-01"] == 5000.0
    assert uacs_map["4-01-02-040-02"] == 3000.0
    assert uacs_map["4-01-02-040-03"] == 800.0
    assert uacs_map["4-01-02-040-04"] == -100.0
    
    assert uacs_map["4-01-02-041-01"] == 5000.0
    assert uacs_map["4-01-02-041-02"] == 3000.0
    assert uacs_map["4-01-02-041-03"] == 800.0
    assert uacs_map["4-01-02-041-04"] == -100.0
